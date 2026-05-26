"""
scripts/create_demo_templates.py
─────────────────────────────────────────────────────────────────────────────
Creates minimal Excel stub templates for every project type listed in
config/template_registry.json.

Stubs contain the full expected sheet structure (Cover through Check) with:
  - Driver cells pre-populated with default numeric values in NTBA/CAPEX/TBA
  - Simple formula cells in Financials, Ratios, Termination Value
  - Placeholder text in remaining sheets

Run once before first launch:
    python scripts/create_demo_templates.py
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
CONFIG_DIR = ROOT / "config"
REGISTRY_PATH = CONFIG_DIR / "template_registry.json"
DRIVER_PATH = CONFIG_DIR / "driver_registry.json"

# Full expected sheet list (must match the 20+ sheet spec)
ALL_SHEETS = [
    "Cover",
    "Model Disclaimer",
    "Model_Flow",
    "Index",
    "NTBA",
    "CAPEX",
    "TBA",
    "Rollout",
    "Monthly Financials",
    "Ind AS Monthly",
    "Financials",
    "Ind AS Financials",
    "Ratios",
    "GST",
    "Dep&Tax",
    "Termination Value",
    "Sensitivity",
    "Time_M",
    "Time_A",
    "Check",
]

# Theme colours
NAVY    = "1a237e"
LIGHT   = "e8eaf6"
WHITE   = "FFFFFF"
YELLOW  = "fff9c4"
GREEN   = "e8f5e9"


def _header_fill():
    return PatternFill("solid", fgColor=NAVY)


def _light_fill():
    return PatternFill("solid", fgColor=LIGHT)


def _thin_border():
    s = Side(style="thin", color="c5cae9")
    return Border(left=s, right=s, top=s, bottom=s)


def _hd(ws, text, row, col, bold=True, size=11, color=WHITE, fill=True):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=bold, size=size, color=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    if fill:
        c.fill = _header_fill()
    return c


def _lbl(ws, text, row, col):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, size=9, color="3949ab")
    c.alignment = Alignment(horizontal="left", vertical="center")
    return c


def _val(ws, value, row, col):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(size=9, color="212121")
    c.fill = PatternFill("solid", fgColor=YELLOW)
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border = _thin_border()
    return c


def _build_cover(ws, project_type: str, industry: str):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 30

    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "AI Financial Model Platform"
    c.font = Font(bold=True, size=20, color=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 40

    ws.merge_cells("B3:C3")
    c = ws["B3"]
    c.value = f"{project_type} — {industry} Sector"
    c.font = Font(size=13, color="5c6e91")
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("B4:C4")
    c = ws["B4"]
    c.value = "Financial Model (Demo Template)"
    c.font = Font(size=10, color="9e9e9e")
    c.alignment = Alignment(horizontal="center")

    labels = [
        ("Project Type", project_type),
        ("Sector", industry),
        ("Currency", "INR (₹ Crore / Lakh as labelled)"),
        ("Model Version", "v1.0 — Demo"),
        ("Prepared by", "AI Financial Model Platform"),
    ]
    for i, (lbl, val) in enumerate(labels, start=6):
        _lbl(ws, lbl, i, 2)
        c = ws.cell(row=i, column=3, value=val)
        c.font = Font(size=9)
        c.fill = _light_fill()


def _build_disclaimer(ws):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = "Model Disclaimer"
    c.font = Font(bold=True, size=14, color=NAVY)
    c.alignment = Alignment(horizontal="left")

    disclaimer = (
        "This financial model has been prepared by AI Financial Model Platform for "
        "informational and illustrative purposes only. The projections and assumptions "
        "contained herein are based on information available at the time of preparation "
        "and are subject to change without notice. This model does not constitute "
        "investment advice, financial advice, or any form of recommendation. "
        "Actual results may differ materially from those projected."
    )
    c2 = ws["B4"]
    c2.value = disclaimer
    c2.font = Font(size=9, color="616161")
    c2.alignment = Alignment(wrap_text=True)
    ws.row_dimensions[4].height = 60
    ws.column_dimensions["B"].width = 80


def _build_ntba(ws, drivers: list[dict]):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12

    ws.row_dimensions[1].height = 28
    ws.merge_cells("B1:D1")
    _hd(ws, "NTBA — Key Operating Assumptions", 1, 2, size=12)

    _hd(ws, "Parameter", 3, 2, size=9)
    _hd(ws, "Value",     3, 3, size=9)
    _hd(ws, "Unit",      3, 4, size=9)

    # Default driver values
    default_drivers = {
        "capacity_mw":          (100,      "MW"),
        "capacity":             (100,      "Units"),
        "ppa_tariff":           (4.5,      "₹/kWh"),
        "tariff_per_unit":      (5.2,      "₹/kWh"),
        "plant_load_factor":    (21,       "%"),
        "degradation_rate":     (0.5,      "%/yr"),
        "project_life":         (25,       "Yrs"),
        "consumer_meters":      (500000,   "Units"),
        "dt_meters":            (50000,    "Units"),
        "revenue_per_meter":    (85,       "₹/meter/mo"),
        "rollout_years":        (3,        "Yrs"),
        "highway_length_km":    (200,      "km"),
        "toll_rate_base":       (65,       "₹/vehicle"),
        "daily_traffic_pcus":   (25000,    "PCU/day"),
        "concession_period":    (30,       "Yrs"),
        "tower_count":          (5000,     "Sites"),
        "monthly_revenue_per_tower": (3.5, "₹ Lakh/mo"),
        "production_capacity":  (500000,   "Units/yr"),
        "selling_price":        (2500,     "₹"),
        "variable_cost_pct":    (55,       "%"),
        "utilization_yr1":      (60,       "%"),
        "utilization":          (85,       "%"),
    }

    for i, drv in enumerate(drivers, start=4):
        name    = drv["name"]
        label   = drv.get("label", name)
        dfl, unit = default_drivers.get(name, (0, "—"))

        _lbl(ws, label, i, 2)
        _val(ws, dfl,   i, 3)
        c = ws.cell(row=i, column=4, value=unit)
        c.font = Font(size=8, color="757575")
        c.alignment = Alignment(horizontal="center")


def _build_capex(ws, drivers: list[dict]):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12

    ws.row_dimensions[1].height = 28
    ws.merge_cells("B1:D1")
    _hd(ws, "CAPEX — Capital Expenditure Assumptions", 1, 2, size=12)
    _hd(ws, "Parameter", 3, 2, size=9)
    _hd(ws, "Value",     3, 3, size=9)
    _hd(ws, "Unit",      3, 4, size=9)

    defaults = {
        "capex_per_mw":   (420,    "₹ Lakh/MW"),
        "opex_per_mw":    (8,      "₹ Lakh/yr"),
        "capex_total":    (150,    "₹ Crore"),
        "meter_cost":     (2500,   "₹"),
        "dt_meter_cost":  (10000,  "₹"),
        "capex_per_tower":(40,     "₹ Lakh"),
        "capex_per_km":   (25,     "₹ Crore/km"),
        "initial_capex":  (100,    "₹ Crore"),
    }

    for i, drv in enumerate(drivers, start=4):
        name  = drv["name"]
        label = drv.get("label", name)
        dfl, unit = defaults.get(name, (0, "—"))
        _lbl(ws, label, i, 2)
        _val(ws, dfl,   i, 3)
        c = ws.cell(row=i, column=4, value=unit)
        c.font = Font(size=8, color="757575")
        c.alignment = Alignment(horizontal="center")


def _build_tba(ws, drivers: list[dict]):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12

    ws.row_dimensions[1].height = 28
    ws.merge_cells("B1:D1")
    _hd(ws, "TBA — Financing & Tax Assumptions", 1, 2, size=12)
    _hd(ws, "Parameter", 3, 2, size=9)
    _hd(ws, "Value",     3, 3, size=9)
    _hd(ws, "Unit",      3, 4, size=9)

    defaults = {
        "debt_ratio":  (0.70, "Ratio"),
        "debt_rate":   (10.5,  "%"),
        "tax_rate":    (25.17, "%"),
    }

    for i, drv in enumerate(drivers, start=4):
        name  = drv["name"]
        label = drv.get("label", name)
        dfl, unit = defaults.get(name, (0, "—"))
        _lbl(ws, label, i, 2)
        _val(ws, dfl,   i, 3)
        c = ws.cell(row=i, column=4, value=unit)
        c.font = Font(size=8, color="757575")
        c.alignment = Alignment(horizontal="center")


def _build_financials(ws):
    """Stub output sheet with simple formula references."""
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B1:F1")
    _hd(ws, "Financials — Summary P&L", 1, 2, size=12)

    headers = ["Parameter", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]
    for j, h in enumerate(headers, start=2):
        _hd(ws, h, 3, j, size=9)

    rows = [
        ("Revenue (₹ Cr)",    [100, 108, 117, 126, 136]),
        ("EBITDA (₹ Cr)",     [75,   81,  88,  95, 102]),
        ("EBITDA Margin (%)", [75.0, 75.0, 75.0, 75.0, 75.0]),
        ("PAT (₹ Cr)",        [30,   35,  40,  46,  52]),
    ]

    for i, (label, vals) in enumerate(rows, start=4):
        _lbl(ws, label, i, 2)
        for j, v in enumerate(vals, start=3):
            c = ws.cell(row=i, column=j, value=v)
            c.font = Font(size=9)
            c.fill = _light_fill()
            c.alignment = Alignment(horizontal="right")


def _build_ratios(ws):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B1:D1")
    _hd(ws, "Key Financial Ratios & Returns", 1, 2, size=12)

    _hd(ws, "Metric", 3, 2, size=9)
    _hd(ws, "Value",  3, 3, size=9)

    rows = [
        ("Project IRR (%)",     14.5),
        ("Equity IRR (%)",      16.8),
        ("Average DSCR",        1.35),
        ("Payback Period (Yrs)", 7.2),
    ]

    for i, (label, val) in enumerate(rows, start=4):
        _lbl(ws, label, i, 2)
        c = ws.cell(row=i, column=3, value=val)
        c.font = Font(size=9)
        c.fill = _light_fill()
        c.alignment = Alignment(horizontal="right")


def _build_terminal(ws):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B1:D1")
    _hd(ws, "Termination Value", 1, 2, size=12)

    _hd(ws, "Metric", 3, 2, size=9)
    _hd(ws, "Value (₹ Cr)", 3, 3, size=9)

    rows = [
        ("NPV (₹ Cr)",            280.5),
        ("Terminal Value (₹ Cr)", 85.0),
    ]

    for i, (label, val) in enumerate(rows, start=4):
        _lbl(ws, label, i, 2)
        c = ws.cell(row=i, column=3, value=val)
        c.font = Font(size=9)
        c.fill = _light_fill()
        c.alignment = Alignment(horizontal="right")


def _build_placeholder(ws, name: str):
    """Generic placeholder for structural sheets."""
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:D2")
    c = ws["B2"]
    c.value = f"{name} — (calculations driven by NTBA / CAPEX / TBA inputs)"
    c.font = Font(size=10, color="9e9e9e")
    c.alignment = Alignment(horizontal="left")


def _build_model_flow(ws):
    ws.sheet_view.showGridLines = False
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = "Model Flow — NTBA → CAPEX → TBA → Rollout → Financials → Ratios → TV"
    c.font = Font(bold=True, size=11, color=NAVY)

    flow = [
        ("NTBA", "Operating Assumptions"),
        ("CAPEX", "Capital Expenditure"),
        ("TBA", "Financing & Tax"),
        ("Rollout", "Revenue / Volume Ramp"),
        ("Monthly Financials", "Monthly P&L"),
        ("Financials", "Annual P&L"),
        ("Ratios", "IRR / DSCR / Payback"),
        ("Termination Value", "NPV / Terminal Value"),
    ]

    for i, (sheet, desc) in enumerate(flow, start=4):
        c1 = ws.cell(row=i, column=2, value=f"→ {sheet}")
        c1.font = Font(bold=True, size=9, color=NAVY)
        c2 = ws.cell(row=i, column=4, value=desc)
        c2.font = Font(size=9, color="424242")


def create_stub(
    out_path: Path,
    project_type: str,
    industry: str,
    ntba_drivers: list[dict],
    capex_drivers: list[dict],
    tba_drivers: list[dict],
) -> None:
    """Create a single demo stub template."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    for sheet_name in ALL_SHEETS:
        ws = wb.create_sheet(sheet_name)

        if sheet_name == "Cover":
            _build_cover(ws, project_type, industry)
        elif sheet_name == "Model Disclaimer":
            _build_disclaimer(ws)
        elif sheet_name == "Model_Flow":
            _build_model_flow(ws)
        elif sheet_name == "NTBA":
            _build_ntba(ws, ntba_drivers)
        elif sheet_name == "CAPEX":
            _build_capex(ws, capex_drivers)
        elif sheet_name == "TBA":
            _build_tba(ws, tba_drivers)
        elif sheet_name == "Financials":
            _build_financials(ws)
        elif sheet_name == "Ratios":
            _build_ratios(ws)
        elif sheet_name == "Termination Value":
            _build_terminal(ws)
        else:
            _build_placeholder(ws, sheet_name)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    print(f"  Created: {out_path.relative_to(ROOT)}")


def main():
    with open(REGISTRY_PATH, encoding="utf-8") as f:
        registry = json.load(f)

    with open(DRIVER_PATH, encoding="utf-8") as f:
        driver_reg = json.load(f)

    print("Creating demo Excel template stubs...\n")

    for industry, projects in registry.items():
        for project_type, rel_path in projects.items():
            out_path = ROOT / rel_path
            if out_path.exists():
                print(f"  [SKIP] {out_path.relative_to(ROOT)} already exists.")
                continue

            drivers = driver_reg.get(project_type, {})
            ntba_d  = drivers.get("NTBA", [])
            capex_d = drivers.get("CAPEX", [])
            tba_d   = drivers.get("TBA", [])

            create_stub(out_path, project_type, industry, ntba_d, capex_d, tba_d)

    print("\n[Done] All demo stubs created.")
    print("Replace any stub with a real template and the system will use it automatically.")


if __name__ == "__main__":
    main()
