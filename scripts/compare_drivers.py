"""
Focused driver extraction: Only scan COLUMNS K-O in NTBA (scenario inputs)
and find all unique labeled numeric cells across all three driver sheets.
Also compare demo.xlsx vs idemo.xlsx values to identify what changed.
"""
import openpyxl
from openpyxl.utils import get_column_letter

DEMO   = "examples/demo.xlsx"
IDEMO  = "examples/idemo.xlsx"

def find_scenario_drivers(path, sheet_name, scenario_col='K'):
    """Find numeric cells in the scenario column with their left-side labels."""
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[sheet_name]
    scol = openpyxl.utils.column_index_from_string(scenario_col)
    results = []

    for row in ws.iter_rows():
        for cell in row:
            if cell.column != scol:
                continue
            if cell.value is None:
                continue
            is_num = isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
            is_formula = isinstance(cell.value, str) and cell.value.strip().startswith('=')
            if not is_num or is_formula:
                continue

            # Find label: search left up to 8 cols
            label = None
            raw_col = None
            for offset in range(1, 9):
                lc = cell.column - offset
                if lc < 1:
                    break
                lv = ws.cell(row=cell.row, column=lc).value
                if isinstance(lv, str) and lv.strip() and not lv.strip().startswith('='):
                    label = lv.strip()
                    raw_col = get_column_letter(lc)
                    break

            addr = f"{scenario_col}{cell.row}"
            results.append({
                'addr': addr,
                'row': cell.row,
                'label': label or f'<Row {cell.row}>',
                'raw_col': raw_col,
                'value': cell.value
            })

    wb.close()
    return results


def find_capex_drivers(path):
    """Find plai numeric input cells in Capex sheet (not in NTBA rollout pattern)."""
    wb = openpyxl.load_workbook(path, read_only=False, data_only=False)
    ws = wb['Capex']
    results = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            is_num = isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
            if not is_num:
                continue
            col_l = get_column_letter(cell.column)
            # Focus on F, G, I, J columns only (unit rate / qty columns)
            if col_l not in ('F', 'G', 'I', 'J'):
                continue

            label = None
            for offset in range(1, 6):
                lc = cell.column - offset
                if lc < 1:
                    break
                lv = ws.cell(row=cell.row, column=lc).value
                if isinstance(lv, str) and lv.strip() and not lv.strip().startswith('='):
                    label = lv.strip()
                    break

            addr = f"{col_l}{cell.row}"
            results.append({
                'addr': addr,
                'row': cell.row,
                'label': label or f'<Row {cell.row}>',
                'col': col_l,
                'value': cell.value
            })
    wb.close()
    return results


print("=" * 80)
print("NTBA SHEET - Scenario K column numeric drivers (DEMO vs IDEMO)")
print("=" * 80)
demo_ntba  = find_scenario_drivers(DEMO,  'NTBA')
idemo_ntba = find_scenario_drivers(IDEMO, 'NTBA')

idemo_map = {r['addr']: r['value'] for r in idemo_ntba}
for r in demo_ntba:
    ival = idemo_map.get(r['addr'], 'N/A')
    changed = '*' if ival != r['value'] else ' '
    print(f"  {changed} {r['addr']:>6}  {repr(r['label']):<55}  demo={r['value']}  idemo={ival}")

print()
print("=" * 80)
print("CAPEX SHEET - Numeric input drivers (DEMO vs IDEMO)")
print("=" * 80)
demo_capex  = find_capex_drivers(DEMO)
idemo_capex = find_capex_drivers(IDEMO)
idemo_map_c = {r['addr']: r['value'] for r in idemo_capex}
for r in demo_capex:
    ival = idemo_map_c.get(r['addr'], 'N/A')
    changed = '*' if ival != r['value'] else ' '
    print(f"  {changed} {r['addr']:>6}  {repr(r['label']):<55}  demo={r['value']}  idemo={ival}")

print()
print("=" * 80)
print("TBA SHEET - Scenario K column numeric drivers (DEMO vs IDEMO)")
print("=" * 80)
demo_tba  = find_scenario_drivers(DEMO,  'TBA')
idemo_tba = find_scenario_drivers(IDEMO, 'TBA')
idemo_map_t = {r['addr']: r['value'] for r in idemo_tba}
for r in demo_tba:
    ival = idemo_map_t.get(r['addr'], 'N/A')
    changed = '*' if ival != r['value'] else ' '
    print(f"  {changed} {r['addr']:>6}  {repr(r['label']):<55}  demo={r['value']}  idemo={ival}")
