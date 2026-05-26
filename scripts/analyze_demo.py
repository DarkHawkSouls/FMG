"""
Analyze demo.xlsx to identify all true driver (numeric, non-formula) cells
and their labels across NTBA, Capex, and TBA sheets.
"""
import openpyxl
from openpyxl.utils import get_column_letter

TEMPLATE = "examples/demo.xlsx"
DRIVER_SHEETS = ["NTBA", "Capex", "TBA"]

wb = openpyxl.load_workbook(TEMPLATE, read_only=False, data_only=False)

print(f"Sheets: {wb.sheetnames}\n")

for sheet_name in DRIVER_SHEETS:
    if sheet_name not in wb.sheetnames:
        print(f"[WARN] Sheet '{sheet_name}' not found")
        continue

    ws = wb[sheet_name]
    print(f"\n{'='*70}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*70}")

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            is_num = isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
            if not is_num:
                continue

            addr = f"{get_column_letter(cell.column)}{cell.row}"

            # Search left for a label (up to 8 columns)
            label = None
            for offset in range(1, 9):
                lc = cell.column - offset
                if lc < 1:
                    break
                lv = ws.cell(row=cell.row, column=lc).value
                if isinstance(lv, str) and lv.strip() and not lv.strip().startswith("="):
                    label = lv.strip()
                    break

            if label is None:
                # Look in col B or C of same row
                for c in [2, 3]:
                    lv = ws.cell(row=cell.row, column=c).value
                    if isinstance(lv, str) and lv.strip():
                        label = lv.strip()
                        break

            if label is None:
                label = f"<Row {cell.row}>"

            print(f"  {addr:>8}  {repr(label):<55}  = {cell.value}")

wb.close()
