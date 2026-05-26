"""
Extract ONLY the meaningful named-label driver cells from NTBA K-column (scenario 1).
This identifies the true user-editable assumption rows.
"""
import openpyxl
from openpyxl.utils import get_column_letter

DEMO = "examples/demo.xlsx"

wb = openpyxl.load_workbook(DEMO, read_only=False, data_only=False)
ws_ntba = wb['NTBA']

print("=== NTBA - All E-column labels with K-column numeric values ===\n")
for row in ws_ntba.iter_rows():
    e_label = None
    k_val = None
    for cell in row:
        col = get_column_letter(cell.column)
        if col == 'E' and isinstance(cell.value, str) and not cell.value.startswith('='):
            e_label = cell.value.strip()
        if col == 'K':
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                k_val = cell.value
    if e_label and k_val is not None:
        print(f"  Row {list(row)[0].row:>4}  K={str(k_val):<15}  NTBA!K{list(row)[0].row}  label={repr(e_label)}")

print("\n\n=== TBA - All E-column labels with K-column numeric values ===\n")
ws_tba = wb['TBA']
for row in ws_tba.iter_rows(max_row=300):
    e_label = None
    k_val = None
    for cell in row:
        col = get_column_letter(cell.column)
        if col == 'E' and isinstance(cell.value, str) and not cell.value.startswith('='):
            e_label = cell.value.strip()
        if col == 'K':
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                k_val = cell.value
    if e_label and k_val is not None:
        print(f"  Row {list(row)[0].row:>4}  K={str(k_val):<15}  TBA!K{list(row)[0].row}  label={repr(e_label)}")

wb.close()
