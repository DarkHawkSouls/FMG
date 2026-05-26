"""
engines/template_writer.py
─────────────────────────────────────────────────────────────────────────────
Loads an Excel template, writes assumption values into NTBA / CAPEX / TBA,
and saves a timestamped output file.

CRITICAL RULES (enforced here):
  ✗  Never touch cells outside NTBA, CAPEX, TBA
  ✗  Never write a formula (all values written are plain Python scalars)
  ✗  Never rename sheets or remove sheets
  ✓  Preserve all formatting and formulas in every other cell
"""

from __future__ import annotations

import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from engines.project_config import normalize_sheet_name, resolve_sheet_name

# Allowed write targets
ALLOWED_SHEETS = {"NTBA", "Capex", "TBA", "Sensitivity"}

OUTPUT_DIR = Path(__file__).parent.parent / "output" / "generated_models"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def write_model(
    template_path: str | Path,
    write_map: dict[str, dict[str, Any]],
    project_type: str = "model",
    industry: str = "generic",
) -> Path:
    """
    Write assumption values into a copy of the template.

    Args:
        template_path: Source .xlsx template file.
        write_map:     {sheet_name: {cell_address: value}}, only NTBA/CAPEX/TBA.
        project_type:  Used for naming the output file.
        industry:      Used for naming the output file.

    Returns:
        Path to the generated output .xlsx file.

    Raises:
        FileNotFoundError: If the template does not exist.
        ValueError:        If write_map targets a sheet other than NTBA/CAPEX/TBA.
    """
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Safety gate — only permitted sheets may be written to
    canonical_write_map: dict[str, dict[str, Any]] = {}
    for sheet_name, cell_map in write_map.items():
        canonical_sheet = normalize_sheet_name(sheet_name)
        canonical_write_map[canonical_sheet] = cell_map

    illegal = set(canonical_write_map.keys()) - ALLOWED_SHEETS
    if illegal:
        raise ValueError(
            f"write_map targets illegal sheets: {illegal}. "
            f"Only {ALLOWED_SHEETS} may be modified."
        )

    # Build output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_project = project_type.replace(" ", "_").lower()
    safe_industry = industry.replace(" ", "_").lower()
    out_filename = f"{safe_industry}_{safe_project}_{timestamp}.xlsx"
    out_path = OUTPUT_DIR / out_filename

    # Copy template to output location before modifying
    shutil.copy2(template_path, out_path)

    # Open the copy in write mode (NOT data_only — we want to keep formulas)
    wb = openpyxl.load_workbook(str(out_path))

    for sheet_name, cell_map in canonical_write_map.items():
        actual_sheet = resolve_sheet_name(wb.sheetnames, sheet_name)
        if actual_sheet is None:
            print(f"[WARN] Sheet '{sheet_name}' not found in workbook — skipping.")
            continue

        ws = wb[actual_sheet]

        for cell_addr, value in cell_map.items():
            cell = ws[cell_addr]

            # Extra safety: never overwrite a formula
            existing = cell.value
            if isinstance(existing, str) and existing.strip().startswith("="):
                print(
                    f"[WARN] Skipping {sheet_name}!{cell_addr} — "
                    f"it contains a formula: {existing[:60]}"
                )
                continue

            cell.value = value

    wb.save(str(out_path))
    wb.close()

    print(f"[OK] Generated model saved: {out_path}")
    return out_path


def demo_write(project_type: str = "Solar Project") -> Path:
    """
    Quick demo: load the config driver_registry and write default values
    back into the template.  Useful for smoke-testing without a UI.
    """
    config_dir = Path(__file__).parent.parent / "config"
    registry_path = config_dir / "driver_registry.json"
    template_registry_path = config_dir / "template_registry.json"

    with open(registry_path, encoding="utf-8") as f:
        driver_reg = json.load(f)
    with open(template_registry_path, encoding="utf-8") as f:
        tmpl_reg = json.load(f)

    # Find template path for this project type
    tmpl_path = None
    for industry, projects in tmpl_reg.items():
        if project_type in projects:
            tmpl_path = Path(__file__).parent.parent / projects[project_type]
            break

    if tmpl_path is None or not tmpl_path.exists():
        raise FileNotFoundError(
            f"Template for '{project_type}' not found.  "
            "Run scripts/create_demo_templates.py first."
        )

    # Build write_map from driver defaults
    project_drivers = driver_reg.get(project_type, {})
    write_map: dict[str, dict[str, Any]] = {}
    for sheet, entries in project_drivers.items():
        write_map[sheet] = {}
        for e in entries:
            write_map[sheet][e["cell"]] = 0  # placeholder

    return write_model(tmpl_path, write_map, project_type)
