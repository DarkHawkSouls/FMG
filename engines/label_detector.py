"""
engines/label_detector.py
─────────────────────────────────────────────────────────────────────────────
Detects human-readable labels for driver candidate cells.

Search strategy (priority order):
  1. Scan up to 10 columns to the LEFT  in the same row  → first text cell wins
  2. Scan up to  5 rows   ABOVE the cell in the same column → first text cell wins
  3. Fallback: "Input_{cell_coordinate}"

After a label is found it is:
  • Cleaned of unit annotations  — "(Rs)", "(%)", "₹", "(MW)", "(Yrs)", etc.
  • Collapsed of extra whitespace
  • Stripped of trailing/leading punctuation

Reads:   config/driver_candidates.json  (or receives dict directly)
Writes:  config/driver_registry.json

Usage (CLI):
    python engines/label_detector.py --template path/to/model.xlsx
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

ROOT       = Path(__file__).parent.parent
CONFIG_DIR = ROOT / "config"

CANDIDATES_PATH = CONFIG_DIR / "driver_candidates.json"
REGISTRY_PATH   = CONFIG_DIR / "driver_registry.json"

# ── How far to search ─────────────────────────────────────────────────────────
MAX_LEFT_COLS  = 10   # columns to search left of the driver cell
MAX_ABOVE_ROWS = 5    # rows to search above the driver cell

# ── Unit / symbol patterns to strip from discovered labels ───────────────────
_UNIT_PATTERNS = [
    r"\(\s*Rs\.?\s*(?:Cr(?:ore)?)?\.?\s*\)",    # (Rs), (Rs Cr), (Rs. Crore)
    r"\(\s*INR\s*\)",
    r"\(\s*₹[^)]*\)",                            # (₹), (₹/kWh), (₹ Lakh) …
    r"\(\s*[A-Za-z%₹][^)]{0,25}\)",              # any short parenthesised unit
    r"[\u20b9$€£]",                              # bare currency symbols
    r"\bINR\b",
    r"\bRs\.?\b",
    r"[-–—]{2,}",                                # long dashes used as separators
]
_UNIT_RE = re.compile("|".join(_UNIT_PATTERNS), re.IGNORECASE)

# Chars to strip from the start/end of a label after unit removal
_EDGE_STRIP = re.compile(r'^[\s\-–—:,./\*#]+|[\s\-–—:,./\*#]+$')


def _is_text_label(value: Any) -> bool:
    """Return True if value is a non-empty, non-formula string."""
    if not isinstance(value, str):
        return False
    s = value.strip()
    return bool(s) and not s.startswith("=")


def clean_label(raw: str) -> str:
    """
    Strip unit annotations and normalise whitespace from a raw cell label.

    Examples:
        "Installed Capacity (MW)"  → "Installed Capacity"
        "Tax Rate (%)"             → "Tax Rate"
        "Capex (Rs. Cr.)"          → "Capex"
        "Revenue ₹ Crore"          → "Revenue  Crore"  → "Revenue Crore"
        "  -- IRR --  "            → "IRR"
    """
    cleaned = _UNIT_RE.sub(" ", raw)
    cleaned = re.sub(r"\s{2,}", " ", cleaned)      # collapse multiple spaces
    cleaned = _EDGE_STRIP.sub("", cleaned)
    return cleaned.strip() or raw.strip()           # never return empty


def _find_label_in_row(ws, row: int, col: int) -> str | None:
    """
    Search left along the row for a text label, up to MAX_LEFT_COLS columns.
    Returns the cleaned label of the first text cell found, or None.
    """
    for offset in range(1, MAX_LEFT_COLS + 1):
        target_col = col - offset
        if target_col < 1:
            break
        val = ws.cell(row=row, column=target_col).value
        if _is_text_label(val):
            return clean_label(str(val))
    return None


def _find_label_above(ws, row: int, col: int) -> str | None:
    """
    Search upward in the same column for a text label, up to MAX_ABOVE_ROWS.
    Returns the cleaned label of the nearest text cell found, or None.
    """
    for offset in range(1, MAX_ABOVE_ROWS + 1):
        target_row = row - offset
        if target_row < 1:
            break
        val = ws.cell(row=target_row, column=col).value
        if _is_text_label(val):
            return clean_label(str(val))
    return None


def detect_labels(
    template_path: str | Path,
    driver_candidates: dict[str, list[dict]] | None = None,
    save_to_config: bool = True,
    config_dir: Path | None = None,
) -> dict[str, list[dict]]:
    """
    Assign human-readable labels to every driver candidate cell.

    Args:
        template_path:     Path to the .xlsx template.
        driver_candidates: Pre-loaded candidates dict.
                           If None, loaded from config/driver_candidates.json.
        save_to_config:    Write driver_registry.json when True.
        config_dir:        Override output directory (for per-session isolation).

    Returns:
        driver_registry := {
            "NTBA":  [
                {
                    "cell":    "C4",
                    "label":   "Installed Capacity",   ← human-readable, units stripped
                    "value":   100,
                    "row":     4,
                    "col":     "C",
                    "raw_label": "Installed Capacity (MW)"   ← original before cleaning
                },
                ...
            ],
            "CAPEX": [...],
            "TBA":   [...],
        }
    """
    template_path = Path(template_path)
    out_dir       = config_dir or CONFIG_DIR

    # Load candidates if not provided
    if driver_candidates is None:
        cand_path = out_dir / "driver_candidates.json"
        if not cand_path.exists():
            raise FileNotFoundError(
                f"driver_candidates.json not found at {cand_path}. "
                "Run template_scanner.py first."
            )
        with open(cand_path, encoding="utf-8") as f:
            driver_candidates = json.load(f)

    # Load template in formula mode so all cell values (not just computed
    # results) are readable — critical for reading text labels next to formulas
    wb = openpyxl.load_workbook(str(template_path), read_only=True, data_only=False)

    driver_registry: dict[str, list[dict]] = {}
    stats = {"found_left": 0, "found_above": 0, "fallback": 0}

    for sheet_name, candidates in driver_candidates.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        sheet_entries: list[dict] = []

        for cand in candidates:
            cell_addr = cand["cell"]
            value     = cand.get("value", 0)

            # Resolve row / col integers
            try:
                cell_obj = ws[cell_addr]
                row = cell_obj.row
                col = cell_obj.column
            except Exception:
                m = re.match(r"([A-Z]+)(\d+)", cell_addr.upper())
                if not m:
                    continue
                col = column_index_from_string(m.group(1))
                row = int(m.group(2))

            raw_label: str | None = None
            source: str           = "fallback"

            # ── Priority 1: search LEFT along the row ──────────────────
            raw_label = _find_label_in_row(ws, row, col)
            if raw_label:
                source = "left"
                stats["found_left"] += 1

            # ── Priority 2: search ABOVE in the same column ────────────
            if not raw_label:
                raw_label = _find_label_above(ws, row, col)
                if raw_label:
                    source = "above"
                    stats["found_above"] += 1

            # ── Fallback ───────────────────────────────────────────────
            if not raw_label:
                raw_label = f"Input_{cell_addr}"
                source    = "fallback"
                stats["fallback"] += 1

            # Clean the label (units, symbols)
            cleaned = clean_label(raw_label) if source != "fallback" else raw_label

            sheet_entries.append({
                "cell":      cell_addr,
                "label":     cleaned,     # ← this is what the UI displays
                "raw_label": raw_label,   # ← original before cleaning (for debug)
                "value":     value,
                "row":       row,
                "col":       get_column_letter(col),
                "source":    source,      # "left" | "above" | "fallback"
            })

        driver_registry[sheet_name] = sheet_entries

    wb.close()

    if save_to_config:
        _write_json(out_dir / "driver_registry.json", driver_registry)
        total = sum(len(v) for v in driver_registry.values())
        print(f"[LabelDetector] {total} drivers labelled  "
              f"(left={stats['found_left']}, above={stats['found_above']}, "
              f"fallback={stats['fallback']})")
        for sheet, entries in driver_registry.items():
            print(f"  {sheet}:")
            for e in entries:
                src_tag = f"[{e['source']}]"
                print(f"    {e['cell']:>6}  {src_tag:<8} "
                      f"label='{e['label']}'  "
                      f"(raw='{e['raw_label']}')")

    return driver_registry


def relabel_curated_registry(
    template_path: str | Path,
    project_type: str,
    registry_path: Path | None = None,
) -> None:
    """
    Re-run label detection for a curated project entry and update
    driver_registry.json in-place.  Useful when replacing a stub
    template with a real client template.
    """
    reg_path   = registry_path or REGISTRY_PATH
    cand_path  = CONFIG_DIR / "driver_candidates.json"

    # First re-scan to get fresh candidates
    from engines.template_scanner import scan_template
    result = scan_template(template_path, save_to_config=False)
    candidates = result["driver_candidates"]

    # Detect labels
    new_entries = detect_labels(
        template_path,
        driver_candidates=candidates,
        save_to_config=False,
    )

    # Merge into existing registry
    if reg_path.exists():
        with open(reg_path, encoding="utf-8") as f:
            reg = json.load(f)
    else:
        reg = {}

    reg[project_type] = new_entries

    with open(reg_path, "w", encoding="utf-8") as f:
        json.dump(reg, f, indent=2, ensure_ascii=False)

    print(f"[LabelDetector] driver_registry.json updated for '{project_type}'")


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)


# ── CLI ───────────────────────────────────────────────────────────────────────
def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(
        description="Detect and clean labels for driver cells in an Excel template."
    )
    parser.add_argument("--template", required=True, help="Path to .xlsx template")
    parser.add_argument(
        "--max-left",
        type=int,
        default=MAX_LEFT_COLS,
        help=f"Max columns to search left (default: {MAX_LEFT_COLS})",
    )
    parser.add_argument(
        "--max-above",
        type=int,
        default=MAX_ABOVE_ROWS,
        help=f"Max rows to search above (default: {MAX_ABOVE_ROWS})",
    )
    args = parser.parse_args()

    # Allow CLI overrides of search depth
    import engines.label_detector as _self
    _self.MAX_LEFT_COLS  = args.max_left
    _self.MAX_ABOVE_ROWS = args.max_above

    detect_labels(args.template)


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    main()
