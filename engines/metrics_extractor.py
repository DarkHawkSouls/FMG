"""
engines/metrics_extractor.py
─────────────────────────────────────────────────────────────────────────────
Reads key financial KPIs from a generated Excel model.

Uses openpyxl with data_only=True to read formula-computed values.

NOTE: Formula results are only available in .xlsx files that have been
saved by Excel (not by openpyxl in write mode).  When using demo stub
templates (created by scripts/create_demo_templates.py), values in output
cells will be 0 or None until a real Excel template is used.

The extractor falls back to a rule-based simulation when cells are None.
"""

from __future__ import annotations

import json
import math
from pathlib import Path
from typing import Any

import openpyxl

# Target sheets and their expected metric cells
METRIC_DEFINITIONS: dict[str, dict[str, dict]] = {
    "Financials": {
        "revenue_cr":       {"cells": ["C5", "D5", "E5"], "label": "Revenue (₹ Cr)", "agg": "first"},
        "ebitda_cr":        {"cells": ["C8", "D8", "E8"], "label": "EBITDA (₹ Cr)",  "agg": "first"},
        "pat_cr":           {"cells": ["C12","D12","E12"],"label": "PAT (₹ Cr)",      "agg": "first"},
        "ebitda_margin_pct":{"cells": ["C9", "D9", "E9"], "label": "EBITDA Margin",   "agg": "first"},
    },
    "Ratios": {
        "irr_pct":          {"cells": ["C5"],              "label": "Project IRR (%)", "agg": "first"},
        "equity_irr_pct":   {"cells": ["C6"],              "label": "Equity IRR (%)",  "agg": "first"},
        "dscr_avg":         {"cells": ["C7"],              "label": "Avg DSCR",        "agg": "first"},
        "payback_yrs":      {"cells": ["C8"],              "label": "Payback (Yrs)",   "agg": "first"},
    },
    "Termination Value": {
        "npv_cr":           {"cells": ["C5"],              "label": "NPV (₹ Cr)",      "agg": "first"},
        "terminal_value_cr":{"cells": ["C6"],              "label": "Terminal Value (₹ Cr)", "agg": "first"},
    },
}


def _safe_float(val) -> float | None:
    """Convert a cell value to float, returning None on failure."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def extract_metrics(model_path: str | Path) -> dict[str, Any]:
    """
    Read key metrics from a generated .xlsx model.

    Args:
        model_path: Path to the generated Excel file.

    Returns:
        Flat dict of {metric_name: value}.
    """
    model_path = Path(model_path)
    if not model_path.exists():
        raise FileNotFoundError(f"Generated model not found: {model_path}")

    wb = openpyxl.load_workbook(str(model_path), read_only=True, data_only=True)
    metrics: dict[str, Any] = {}

    for sheet_name, metric_map in METRIC_DEFINITIONS.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        for metric_key, meta in metric_map.items():
            for cell_addr in meta["cells"]:
                val = _safe_float(ws[cell_addr].value)
                if val is not None:
                    metrics[metric_key] = val
                    break

    wb.close()
    return metrics


def simulate_metrics(assumptions: dict[str, Any], project_type: str) -> dict[str, Any]:
    """
    Rule-based metric simulation for demo mode when real Excel formulas
    can't be evaluated (e.g., stub templates or fresh openpyxl saves).

    Provides realistic-looking but approximate KPIs.
    """
    metrics: dict[str, Any] = {}

    # ── Solar / Wind ──────────────────────────────────────────────────────
    if project_type in ("Solar Project", "Wind Project"):
        capacity   = float(assumptions.get("capacity_mw", 100))
        tariff     = float(assumptions.get("ppa_tariff", assumptions.get("tariff_per_unit", 4.5)))
        plf        = float(assumptions.get("plant_load_factor", 21)) / 100
        capex_lakh = float(assumptions.get("capex_per_mw", 420))
        debt_ratio = float(assumptions.get("debt_ratio", 0.70))
        tax_rate   = float(assumptions.get("tax_rate", 25.17)) / 100
        life       = int(assumptions.get("project_life", 25))

        annual_units_mn = capacity * plf * 8760 / 1_000_000  # MU
        revenue_cr = annual_units_mn * 10 * tariff / 100      # ₹ Crore (1 MU = 10Lakh kWh)
        total_capex_cr = capex_lakh * capacity / 100

        ebitda_margin = 0.75
        ebitda_cr = revenue_cr * ebitda_margin
        dep_cr = total_capex_cr / life
        interest_cr = total_capex_cr * debt_ratio * 0.105
        pbt = ebitda_cr - dep_cr - interest_cr
        pat_cr = max(0, pbt * (1 - tax_rate))

        # Simplified IRR approximation
        equity_cr = total_capex_cr * (1 - debt_ratio)
        irr_pct = round((pat_cr / equity_cr) * 100, 2) if equity_cr > 0 else 0
        dscr = (ebitda_cr / max(0.01, interest_cr + total_capex_cr * debt_ratio / 15))
        payback = equity_cr / max(0.01, pat_cr)

        # NPV (simplified)
        wacc = 0.12
        npv_cr = sum(
            pat_cr / ((1 + wacc) ** yr) for yr in range(1, life + 1)
        ) - equity_cr

        metrics = {
            "revenue_cr":        round(revenue_cr, 2),
            "ebitda_cr":         round(ebitda_cr,  2),
            "pat_cr":            round(pat_cr,      2),
            "ebitda_margin_pct": round(ebitda_margin * 100, 1),
            "irr_pct":           round(irr_pct,     2),
            "equity_irr_pct":    round(irr_pct * 0.85, 2),
            "dscr_avg":          round(dscr,         2),
            "payback_yrs":       round(payback,       1),
            "npv_cr":            round(npv_cr,        2),
            "terminal_value_cr": round(total_capex_cr * 0.3, 2),
        }

    # ── Smart Meter ───────────────────────────────────────────────────────
    elif project_type == "Smart Meter Rollout":
        meters   = int(assumptions.get("consumer_meters", 500000))
        rev_pm   = float(assumptions.get("revenue_per_meter", 85))
        m_cost   = float(assumptions.get("meter_cost", 2500))
        d_ratio  = float(assumptions.get("debt_ratio", 0.75))
        tax_rate = float(assumptions.get("tax_rate", 25.17)) / 100

        revenue_cr = meters * rev_pm * 12 / 1e7
        total_capex_cr = meters * m_cost / 1e7
        ebitda_cr = revenue_cr * 0.60
        dep_cr = total_capex_cr / 10
        interest_cr = total_capex_cr * d_ratio * 0.105
        pat_cr = max(0, (ebitda_cr - dep_cr - interest_cr) * (1 - tax_rate))
        equity_cr = total_capex_cr * (1 - d_ratio)
        irr_pct = round((pat_cr / max(0.01, equity_cr)) * 100, 2)

        metrics = {
            "revenue_cr":        round(revenue_cr, 2),
            "ebitda_cr":         round(ebitda_cr,  2),
            "pat_cr":            round(pat_cr,      2),
            "ebitda_margin_pct": 60.0,
            "irr_pct":           irr_pct,
            "equity_irr_pct":    round(irr_pct * 0.80, 2),
            "dscr_avg":          round(ebitda_cr / max(0.01, interest_cr), 2),
            "payback_yrs":       round(equity_cr / max(0.01, pat_cr), 1),
            "npv_cr":            round(pat_cr * 6 - equity_cr, 2),
            "terminal_value_cr": round(total_capex_cr * 0.2, 2),
        }

    # ── Highway PPP ───────────────────────────────────────────────────────
    elif project_type == "Highway PPP":
        length      = float(assumptions.get("highway_length_km", 200))
        toll        = float(assumptions.get("toll_rate_base", 65))
        traffic     = int(assumptions.get("daily_traffic_pcus", 25000))
        capex_per_km= float(assumptions.get("capex_per_km", 25))
        d_ratio     = float(assumptions.get("debt_ratio", 0.80))
        concession  = int(assumptions.get("concession_period", 30))
        tax_rate    = float(assumptions.get("tax_rate", 25.17)) / 100

        revenue_cr  = traffic * toll * 365 * length / 1e7
        capex_cr    = length * capex_per_km
        ebitda_cr   = revenue_cr * 0.65
        dep_cr      = capex_cr / concession
        interest_cr = capex_cr * d_ratio * 0.095
        pat_cr      = max(0, (ebitda_cr - dep_cr - interest_cr) * (1 - tax_rate))
        equity_cr   = capex_cr * (1 - d_ratio)
        irr_pct     = round((pat_cr / max(0.01, equity_cr)) * 100, 2)

        metrics = {
            "revenue_cr":        round(revenue_cr, 2),
            "ebitda_cr":         round(ebitda_cr,  2),
            "pat_cr":            round(pat_cr,      2),
            "ebitda_margin_pct": 65.0,
            "irr_pct":           irr_pct,
            "equity_irr_pct":    round(irr_pct * 0.85, 2),
            "dscr_avg":          round(ebitda_cr / max(0.01, interest_cr), 2),
            "payback_yrs":       round(equity_cr / max(0.01, pat_cr), 1),
            "npv_cr":            round(pat_cr * 8 - equity_cr, 2),
            "terminal_value_cr": round(capex_cr * 0.15, 2),
        }

    # ── Network Rollout ───────────────────────────────────────────────────
    elif project_type == "Network Rollout":
        towers   = int(assumptions.get("tower_count", 5000))
        rev_pm   = float(assumptions.get("monthly_revenue_per_tower", 3.5))
        capex_t  = float(assumptions.get("capex_per_tower", 40))
        d_ratio  = float(assumptions.get("debt_ratio", 0.65))
        tax_rate = float(assumptions.get("tax_rate", 25.17)) / 100

        revenue_cr  = towers * rev_pm * 12 / 100
        capex_cr    = towers * capex_t / 100
        ebitda_cr   = revenue_cr * 0.55
        dep_cr      = capex_cr / 15
        interest_cr = capex_cr * d_ratio * 0.10
        pat_cr      = max(0, (ebitda_cr - dep_cr - interest_cr) * (1 - tax_rate))
        equity_cr   = capex_cr * (1 - d_ratio)
        irr_pct     = round((pat_cr / max(0.01, equity_cr)) * 100, 2)

        metrics = {
            "revenue_cr":        round(revenue_cr, 2),
            "ebitda_cr":         round(ebitda_cr,  2),
            "pat_cr":            round(pat_cr,      2),
            "ebitda_margin_pct": 55.0,
            "irr_pct":           irr_pct,
            "equity_irr_pct":    round(irr_pct * 0.85, 2),
            "dscr_avg":          round(ebitda_cr / max(0.01, interest_cr), 2),
            "payback_yrs":       round(equity_cr / max(0.01, pat_cr), 1),
            "npv_cr":            round(pat_cr * 7 - equity_cr, 2),
            "terminal_value_cr": round(capex_cr * 0.2, 2),
        }

    # ── Plant Expansion ───────────────────────────────────────────────────
    elif project_type == "Plant Expansion":
        capacity   = float(assumptions.get("production_capacity", 500000))
        price      = float(assumptions.get("selling_price", 2500))
        var_cost_p = float(assumptions.get("variable_cost_pct", 55)) / 100
        util_yr1   = float(assumptions.get("utilization_yr1", 60)) / 100
        capex_cr   = float(assumptions.get("capex_total", 150))
        d_ratio    = float(assumptions.get("debt_ratio", 0.65))
        life       = int(assumptions.get("project_life", 15))
        tax_rate   = float(assumptions.get("tax_rate", 25.17)) / 100

        revenue_cr  = capacity * util_yr1 * price / 1e7
        ebitda_cr   = revenue_cr * (1 - var_cost_p) * 0.85
        dep_cr      = capex_cr / life
        interest_cr = capex_cr * d_ratio * 0.10
        pat_cr      = max(0, (ebitda_cr - dep_cr - interest_cr) * (1 - tax_rate))
        equity_cr   = capex_cr * (1 - d_ratio)
        irr_pct     = round((pat_cr / max(0.01, equity_cr)) * 100, 2)

        metrics = {
            "revenue_cr":        round(revenue_cr, 2),
            "ebitda_cr":         round(ebitda_cr,  2),
            "pat_cr":            round(pat_cr,      2),
            "ebitda_margin_pct": round((1 - var_cost_p) * 85, 1),
            "irr_pct":           irr_pct,
            "equity_irr_pct":    round(irr_pct * 0.85, 2),
            "dscr_avg":          round(ebitda_cr / max(0.01, interest_cr), 2),
            "payback_yrs":       round(equity_cr / max(0.01, pat_cr), 1),
            "npv_cr":            round(pat_cr * 6 - equity_cr, 2),
            "terminal_value_cr": round(capex_cr * 0.25, 2),
        }

    else:
        metrics = {
            "revenue_cr": 0, "ebitda_cr": 0, "pat_cr": 0,
            "ebitda_margin_pct": 0, "irr_pct": 0, "equity_irr_pct": 0,
            "dscr_avg": 0, "payback_yrs": 0, "npv_cr": 0, "terminal_value_cr": 0,
        }

    return metrics


METRIC_LABELS: dict[str, str] = {
    "revenue_cr":        "Revenue (₹ Crore)",
    "ebitda_cr":         "EBITDA (₹ Crore)",
    "pat_cr":            "PAT (₹ Crore)",
    "ebitda_margin_pct": "EBITDA Margin (%)",
    "irr_pct":           "Project IRR (%)",
    "equity_irr_pct":    "Equity IRR (%)",
    "dscr_avg":          "Average DSCR",
    "payback_yrs":       "Payback Period (Yrs)",
    "npv_cr":            "NPV (₹ Crore)",
    "terminal_value_cr": "Terminal Value (₹ Crore)",
}
