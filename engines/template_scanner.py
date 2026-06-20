"""
engines/template_scanner.py
─────────────────────────────────────────────────────────────────────────────
Reverse-engineers an uploaded Excel template workbook.

Outputs three JSON files to config/:
  • sheet_structure.json    — all sheets, their dimensions, type classification
  • formula_registry.json  — every formula cell across the workbook
  • driver_candidates.json — numeric, non-formula cells in NTBA / CAPEX / TBA

Usage (CLI):
    python engines/template_scanner.py --template path/to/model.xlsx
"""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

import os

ROOT       = Path(__file__).parent.parent
CONFIG_DIR = ROOT / "config"
if not os.environ.get("VERCEL"):
    CONFIG_DIR.mkdir(parents=True, exist_ok=True)

# Sheets where driver inputs live — only these are scanned for candidates
DRIVER_SHEETS = ["NTBA", "Capex", "TBA", "Sensitivity"]

# Sheet classifications
CALC_KEYWORDS    = {"financial", "ratio", "p&l", "monthly", "rollout", "dep", "gst", "sensitivity", "check", "time"}
INPUT_KEYWORDS   = {"ntba", "capex", "tba", "index", "cover"}
SUMMARY_KEYWORDS = {"termination", "summary", "output"}


def _classify_sheet(name: str) -> str:
    low = name.lower()
    if any(k in low for k in INPUT_KEYWORDS):
        return "input"
    if any(k in low for k in SUMMARY_KEYWORDS):
        return "summary"
    if any(k in low for k in CALC_KEYWORDS):
        return "calculation"
    return "other"


def _is_formula(value: Any) -> bool:
    return isinstance(value, str) and value.strip().startswith("=")


def _is_numeric(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def scan_template(
    template_path: str | Path,
    driver_sheets: list[str] | None = None,
    save_to_config: bool = True,
    config_dir: Path | None = None,
) -> dict[str, Any]:
    """
    Full workbook scan.

    Args:
        template_path:  Path to the .xlsx template.
        driver_sheets:  Sheets to scan for driver candidates (default: NTBA, CAPEX, TBA).
        save_to_config: Write JSON output files to config/.
        config_dir:     Override config directory (useful for per-session isolation).

    Returns:
        {
            "sheet_structure":    {...},
            "formula_registry":   {...},
            "driver_candidates":  {...},
        }
    """
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    target_sheets = driver_sheets or DRIVER_SHEETS
    out_dir       = config_dir or CONFIG_DIR

    # ── Load workbook in formula mode ──────────────────────────────────────
    wb = openpyxl.load_workbook(str(template_path), read_only=True, data_only=False)

    sheet_structure:   dict[str, Any] = {}
    formula_registry:  dict[str, list] = {}
    driver_candidates: dict[str, list] = {}

    total_formula_cells  = 0
    total_numeric_cells  = 0
    total_driver_cells   = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Collect dimensions safely (read_only mode has min/max row/col)
        try:
            min_row = ws.min_row or 1
            max_row = ws.max_row or 1
            min_col = ws.min_column or 1
            max_col = ws.max_column or 1
        except Exception:
            min_row = max_row = min_col = max_col = 1

        sheet_formula_cells  = []
        sheet_numeric_cells  = []
        sheet_driver_cells   = []

        is_driver_sheet = sheet_name in target_sheets

        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if val is None:
                    continue

                col_letter = get_column_letter(cell.column)
                addr = f"{col_letter}{cell.row}"

                if _is_formula(val):
                    sheet_formula_cells.append({
                        "cell":    addr,
                        "formula": str(val)[:200],   # truncate very long formulas
                        "row":     cell.row,
                        "col":     col_letter,
                    })
                    total_formula_cells += 1

                elif _is_numeric(val):
                    sheet_numeric_cells.append({
                        "cell":  addr,
                        "value": val,
                        "row":   cell.row,
                        "col":   col_letter,
                    })
                    total_numeric_cells += 1

                    if is_driver_sheet:
                        sheet_driver_cells.append({
                            "cell":  addr,
                            "value": val,
                            "row":   cell.row,
                            "col":   col_letter,
                        })
                        total_driver_cells += 1

        # Sheet structure
        sheet_structure[sheet_name] = {
            "index":          wb.sheetnames.index(sheet_name),
            "classification": _classify_sheet(sheet_name),
            "dimensions": {
                "min_row": min_row, "max_row": max_row,
                "min_col": min_col, "max_col": max_col,
            },
            "formula_cell_count":  len(sheet_formula_cells),
            "numeric_cell_count":  len(sheet_numeric_cells),
            "is_driver_sheet":     is_driver_sheet,
        }

        formula_registry[sheet_name]  = sheet_formula_cells

        if is_driver_sheet:
            driver_candidates[sheet_name] = sheet_driver_cells

    wb.close()

    summary = {
        "template_path":       str(template_path),
        "sheet_count":         len(wb.sheetnames),
        "total_formula_cells": total_formula_cells,
        "total_numeric_cells": total_numeric_cells,
        "total_driver_candidates": total_driver_cells,
        "driver_sheets_scanned":   target_sheets,
    }

    result = {
        "sheet_structure":   sheet_structure,
        "formula_registry":  formula_registry,
        "driver_candidates": driver_candidates,
        "summary":           summary,
    }

    if save_to_config:
        _write_json(out_dir / "sheet_structure.json",   sheet_structure)
        _write_json(out_dir / "formula_registry.json",  formula_registry)
        _write_json(out_dir / "driver_candidates.json", driver_candidates)
        print(f"[Scanner] sheet_structure.json   — {len(sheet_structure)} sheets")
        print(f"[Scanner] formula_registry.json  — {total_formula_cells} formula cells")
        print(f"[Scanner] driver_candidates.json — {total_driver_cells} driver candidates")

    return result


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(description="Scan an Excel template.")
    parser.add_argument("--template", required=True)
    parser.add_argument(
        "--sheets",
        nargs="+",
        default=DRIVER_SHEETS,
        help="Driver input sheets to scan (default: NTBA CAPEX TBA)",
    )
    args = parser.parse_args()

    result = scan_template(args.template, driver_sheets=args.sheets)
    s = result["summary"]
    print(f"\nScan complete: {s['sheet_count']} sheets, "
          f"{s['total_formula_cells']} formulas, "
          f"{s['total_driver_candidates']} driver candidates")


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    main()
