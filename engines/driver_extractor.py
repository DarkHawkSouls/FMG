"""
engines/driver_extractor.py
─────────────────────────────────────────────────────────────────────────────
Scans NTBA / CAPEX / TBA sheets of an Excel template and detects numeric,
non-formula cells that are likely driver inputs.

Usage (CLI):
    python engines/driver_extractor.py \
        --template templates/energy/solar_project_model.xlsx \
        --project "Solar Project"

Output updates config/driver_registry.json with the detected cells.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

# Sheets that may contain user-editable driver inputs
DRIVER_SHEETS = ["NTBA", "Capex", "TBA", "Sensitivity"]

CONFIG_DIR = Path(__file__).parent.parent / "config"
REGISTRY_PATH = CONFIG_DIR / "driver_registry.json"


def _is_formula(value) -> bool:
    """Return True if the cell value looks like a formula string."""
    return isinstance(value, str) and value.strip().startswith("=")


def _is_numeric_value(value) -> bool:
    """Return True if the cell holds a plain numeric value."""
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def extract_drivers(template_path: str | Path) -> dict[str, list[dict]]:
    """
    Open a template (formula mode) and detect driver cells in the three
    designated sheets.

    Returns:
        {
            "NTBA":  [{"cell": "C8", "name": "row_label_or_addr", "label": ...}, ...],
            "CAPEX": [...],
            "TBA":   [...],
        }
    """
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    # Load in formula mode (data_only=False so we can detect formula cells)
    wb = openpyxl.load_workbook(str(template_path), read_only=True, data_only=False)

    drivers: dict[str, list[dict]] = {sheet: [] for sheet in DRIVER_SHEETS}

    for sheet_name in DRIVER_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  [WARN] Sheet '{sheet_name}' not found in template — skipping.")
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                raw = cell.value
                # Skip blanks, formulas, strings
                if raw is None or _is_formula(raw) or not _is_numeric_value(raw):
                    continue

                addr = f"{get_column_letter(cell.column)}{cell.row}"

                # Label resolution priority:
                #   1. Cell immediately to the LEFT  (col - 1)  — covers B/C layout used by most templates
                #   2. Column A of the same row — covers A/B layout used by some templates
                #   3. Fall back to cell address
                label = None

                if cell.column > 1:
                    left_val = ws.cell(row=cell.row, column=cell.column - 1).value
                    if left_val and isinstance(left_val, str) and not left_val.strip().startswith("="):
                        label = left_val.strip()

                if not label:
                    col_a_val = ws.cell(row=cell.row, column=1).value
                    if col_a_val and isinstance(col_a_val, str) and not col_a_val.strip().startswith("="):
                        label = col_a_val.strip()

                if not label:
                    label = addr

                # Derive a snake_case name from the label
                name = re.sub(r"[^a-z0-9]+", "_", label.lower()).strip("_") or addr.lower()

                drivers[sheet_name].append(
                    {
                        "cell":  addr,
                        "name":  name,
                        "label": label,
                        "value": raw,  # original default value from template
                    }
                )

    wb.close()
    return drivers


def _align_with_schema(
    project_type: str,
    drivers: dict[str, list[dict]],
) -> dict[str, list[dict]]:
    """
    Re-map driver 'name' fields to match input_schema.json field names by
    position order within each sheet.  This ensures assumption_mapper resolves
    correctly even when template labels differ from schema field names.

    Alignment strategy:
      - For each sheet (NTBA / CAPEX / TBA), take the ordered list of detected
        driver cells.
      - Match them position-by-position against the schema fields that belong
        to this sheet according to driver_registry defaults.
      - Any extra detected cells beyond the schema count keep their raw names.
    """
    schema_path = CONFIG_DIR / "input_schema.json"
    seed_registry_path = CONFIG_DIR / "driver_registry.json"

    if not schema_path.exists() or not seed_registry_path.exists():
        return drivers  # can't align without schema

    with open(schema_path, encoding="utf-8") as f:
        input_schema = json.load(f)
    with open(seed_registry_path, encoding="utf-8") as f:
        seed_registry = json.load(f)

    project_schema_fields = [f["name"] for f in input_schema.get(project_type, [])]
    seed_entries = seed_registry.get(project_type, {})

    # Build ordered list: sheet → [name, name, ...] from seed
    sheet_field_order: dict[str, list[str]] = {}
    for sheet_name, entries in seed_entries.items():
        sheet_field_order[sheet_name] = [e["name"] for e in entries]

    aligned: dict[str, list[dict]] = {}
    for sheet_name, cell_list in drivers.items():
        ordered_names = sheet_field_order.get(sheet_name, [])
        new_list = []
        for i, cell_info in enumerate(cell_list):
            if i < len(ordered_names):
                # Substitute the canonical schema name but keep extracted label
                new_list.append({
                    "cell":  cell_info["cell"],
                    "name":  ordered_names[i],
                    "label": cell_info["label"],
                    "value": cell_info.get("value"),
                })
            else:
                new_list.append(cell_info)
        aligned[sheet_name] = new_list

    return aligned


def update_registry(
    project_type: str,
    drivers: dict[str, list[dict]],
    align_schema: bool = True,
) -> None:
    """Merge detected drivers into the global driver_registry.json."""
    if align_schema:
        drivers = _align_with_schema(project_type, drivers)

    registry: dict = {}
    if REGISTRY_PATH.exists():
        with open(REGISTRY_PATH, encoding="utf-8") as f:
            registry = json.load(f)

    # Strip the 'value' key — we only persist cell/name/label
    clean: dict[str, list] = {}
    for sheet, cells in drivers.items():
        clean[sheet] = [
            {"cell": c["cell"], "name": c["name"], "label": c["label"]}
            for c in cells
        ]

    registry[project_type] = clean

    with open(REGISTRY_PATH, "w", encoding="utf-8") as f:
        json.dump(registry, f, indent=2, ensure_ascii=False)

    print(f"[OK] driver_registry.json updated for '{project_type}'")


# ── CLI entry point ──────────────────────────────────────────────────────────

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract driver cells from an Excel financial model template."
    )
    parser.add_argument("--template",      required=True,  help="Path to the .xlsx template")
    parser.add_argument("--project",       required=True,  help='Project type name e.g. "Solar Project"')
    parser.add_argument("--no-align",      action="store_true", help="Skip schema-name alignment step")
    args = parser.parse_args()

    print(f"Scanning template : {args.template}")
    print(f"Project type      : {args.project}")
    print()

    drivers = extract_drivers(args.template)

    total = sum(len(v) for v in drivers.values())
    print(f"Found {total} driver cells across {len(DRIVER_SHEETS)} sheets (before alignment):")
    for sheet, cells in drivers.items():
        print(f"  {sheet}: {len(cells)} cells")
        for c in cells:
            print(f"    {c['cell']:>6}  {c['name']:<40}  label='{c['label']}'  default={c['value']}")

    align = not args.no_align
    update_registry(args.project, drivers, align_schema=align)

    if align:
        # Re-load to show final aligned names
        with open(REGISTRY_PATH, encoding="utf-8") as f:
            final = json.load(f)
        project_final = final.get(args.project, {})
        print()
        print("Final aligned driver registry:")
        for sheet, entries in project_final.items():
            print(f"  {sheet}:")
            for e in entries:
                print(f"    {e['cell']:>6}  {e['name']:<40}  label='{e['label']}'")


if __name__ == "__main__":
    main()
