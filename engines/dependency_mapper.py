"""
engines/dependency_mapper.py
─────────────────────────────────────────────────────────────────────────────
Determines sheet-level dependency relationships by parsing formula cross-
references in the workbook.

Output: config/sheet_dependencies.json

Example output:
{
  "NTBA":              ["Rollout", "Monthly Financials"],
  "CAPEX":             ["Monthly Financials", "Dep&Tax"],
  "TBA":               ["Monthly Financials"],
  "Rollout":           ["Monthly Financials"],
  "Monthly Financials":["Financials", "Ind AS Monthly"],
  "Financials":        ["Ratios", "Termination Value", "Sensitivity"]
}

Usage:
    python engines/dependency_mapper.py --template path/to/model.xlsx
"""

from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

import openpyxl

ROOT       = Path(__file__).parent.parent
CONFIG_DIR = ROOT / "config"

# Regex to find external sheet references inside formulas:
# Pattern: SheetName!CellRef  or  'Sheet Name'!CellRef
SHEET_REF_PATTERN = re.compile(
    r"(?:'([^']+)'|([A-Za-z0-9_\-\.]+))!"
)


def extract_sheet_references(formula: str) -> list[str]:
    """Return all sheet names referenced in a formula string."""
    refs = []
    for match in SHEET_REF_PATTERN.finditer(formula):
        sheet_name = match.group(1) or match.group(2)
        if sheet_name:
            refs.append(sheet_name.strip())
    return refs


def map_dependencies(
    template_path: str | Path,
    save_to_config: bool = True,
    config_dir: Path | None = None,
) -> dict[str, list[str]]:
    """
    Parse all formula cells to build a sheet→sheets dependency graph.

    Returns:
        {source_sheet: [dependent_sheet, ...], ...}
        Interpretation: source_sheet formulas reference data from dependent_sheets.
        (i.e. source_sheet DEPENDS ON dependent_sheets)
    """
    out_dir = config_dir or CONFIG_DIR
    template_path = Path(template_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    wb = openpyxl.load_workbook(str(template_path), read_only=True, data_only=False)

    sheet_names = set(wb.sheetnames)

    # {sheet_name: set of sheets it references}
    dependencies: defaultdict[str, set[str]] = defaultdict(set)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not isinstance(val, str) or not val.strip().startswith("="):
                    continue
                refs = extract_sheet_references(val)
                for ref in refs:
                    if ref in sheet_names and ref != sheet_name:
                        dependencies[sheet_name].add(ref)

    wb.close()

    # Convert sets to sorted lists for JSON serialisation
    deps_list: dict[str, list[str]] = {
        sheet: sorted(refs)
        for sheet, refs in dependencies.items()
    }

    # Add sheets that have no outgoing deps (they exist but reference nothing)
    for sheet in sheet_names:
        if sheet not in deps_list:
            deps_list[sheet] = []

    if save_to_config:
        _write_json(out_dir / "sheet_dependencies.json", deps_list)
        total_edges = sum(len(v) for v in deps_list.values())
        print(f"[DependencyMapper] sheet_dependencies.json — {total_edges} dependency edges")
        # Print non-empty entries
        for src, dests in sorted(deps_list.items()):
            if dests:
                print(f"  {src} → {', '.join(dests)}")

    return deps_list


def build_topological_order(deps: dict[str, list[str]]) -> list[str]:
    """
    Return sheets in topological order (most upstream first).
    Uses Kahn's algorithm; handles cycles by leaving them in arbitrary order.
    """
    try:
        import networkx as nx  # type: ignore
        G = nx.DiGraph()
        for sheet, refs in deps.items():
            G.add_node(sheet)
            for ref in refs:
                G.add_edge(ref, sheet)  # ref feeds into sheet
        try:
            return list(nx.topological_sort(G))
        except nx.NetworkXUnfeasible:
            return list(G.nodes)
    except ImportError:
        # Fallback without networkx
        return list(deps.keys())


def _write_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False, default=str)


# ── CLI ───────────────────────────────────────────────────────────────────────

def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(
        description="Map sheet dependencies from formula cross-references."
    )
    parser.add_argument("--template", required=True)
    args = parser.parse_args()
    map_dependencies(args.template)


if __name__ == "__main__":
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    main()
