"""
webapp/app.py  —  AI Financial Model Platform v2.0
─────────────────────────────────────────────────────────────────────────────
Tab 1 — Curated Templates  : pre-configured industry models
Tab 2 — Upload Template    : upload ANY Excel → scan → dynamic form → generate
"""
from __future__ import annotations

import io
import json
import os
import shutil
import sys
import tempfile
from pathlib import Path
from typing import Any

import streamlit as st

from engines.project_config import (
    flatten_project_fields,
    load_driver_registry as load_driver_registry_config,
    load_input_schema as load_input_schema_config,
)

ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

# ── Page config (must be first Streamlit call) ────────────────────────────────
st.set_page_config(
    page_title="AI Financial Model Platform",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}

.stApp{background:linear-gradient(135deg,#0d1117 0%,#0d1b3e 50%,#091529 100%);color:#e8eaf6;}

[data-testid="stSidebar"]{
    background:linear-gradient(180deg,#0a1628 0%,#0d1f45 100%);
    border-right:1px solid #1a2f5c;}
[data-testid="stSidebar"] .stMarkdown,
[data-testid="stSidebar"] label{color:#c5cae9 !important;}

.platform-banner{
    background:linear-gradient(135deg,#0d1b3e 0%,#1a237e 50%,#0d1b3e 100%);
    border:1px solid #283593;border-radius:14px;padding:24px;
    text-align:center;margin-bottom:20px;}
.platform-banner h1{
    font-size:1.9rem;font-weight:700;
    background:linear-gradient(135deg,#7986cb,#ffffff,#9fa8da);
    -webkit-background-clip:text;-webkit-text-fill-color:transparent;margin:0;}
.platform-banner p{color:#9fa8da;font-size:.88rem;margin-top:4px;}

.step-header{
    background:linear-gradient(90deg,#1a237e 0%,#283593 100%);
    border-radius:10px;padding:10px 18px;margin-bottom:14px;border-left:4px solid #7986cb;}
.step-header h3{color:#fff;margin:0;font-size:1rem;font-weight:600;}
.step-header p{color:#9fa8da;margin:2px 0 0;font-size:.8rem;}

.kpi-tile{
    background:linear-gradient(135deg,#1a237e 0%,#283593 100%);
    border-radius:10px;padding:14px;text-align:center;border:1px solid #3949ab;margin:4px;}
.kpi-tile .label{font-size:.7rem;color:#9fa8da;text-transform:uppercase;letter-spacing:.05em;margin-bottom:5px;}
.kpi-tile .value{font-size:1.3rem;font-weight:700;color:#fff;}
.kpi-tile .signal{font-size:1.1rem;}

.metric-card{
    background:linear-gradient(135deg,#0d1f45 0%,#1a237e22 100%);
    border:1px solid #283593;border-radius:12px;padding:16px 20px;margin:5px 0;
    transition:transform .2s,box-shadow .2s;}
.metric-card:hover{transform:translateY(-2px);box-shadow:0 8px 32px rgba(26,35,126,.3);}

.scanner-box{
    background:#0a1628;border:1px solid #1a2f5c;border-radius:10px;padding:16px;margin:6px 0;}

.stButton > button{
    background:linear-gradient(135deg,#1a237e 0%,#3949ab 100%);
    color:white;border:none;border-radius:8px;padding:10px 24px;
    font-weight:600;font-size:.88rem;width:100%;transition:all .2s;}
.stButton > button:hover{
    background:linear-gradient(135deg,#283593 0%,#5c6bc0 100%);
    transform:translateY(-1px);box-shadow:0 4px 20px rgba(57,73,171,.4);}
.stDownloadButton > button{
    background:linear-gradient(135deg,#1b5e20 0%,#2e7d32 100%);
    color:white;border:none;border-radius:8px;width:100%;}
.stDownloadButton > button:hover{
    background:linear-gradient(135deg,#2e7d32 0%,#43a047 100%);}

.stTabs [data-baseweb="tab-list"]{background:#0a1628;border-radius:8px;gap:4px;}
.stTabs [data-baseweb="tab"]{background:transparent;color:#9fa8da;border-radius:6px;}
.stTabs [aria-selected="true"]{background:#1a237e !important;color:white !important;}

.alert-success{background:#052e16;border-left:4px solid #22c55e;border-radius:6px;
    padding:10px 14px;color:#86efac;font-size:.85rem;margin:6px 0;}
.alert-warning{background:#332200;border-left:4px solid #f59e0b;border-radius:6px;
    padding:10px 14px;color:#fde68a;font-size:.85rem;margin:6px 0;}
.alert-info{background:#0c1a3a;border-left:4px solid #3b82f6;border-radius:6px;
    padding:10px 14px;color:#93c5fd;font-size:.85rem;margin:6px 0;}
.alert-error{background:#2d0a0a;border-left:4px solid #ef4444;border-radius:6px;
    padding:10px 14px;color:#fca5a5;font-size:.85rem;margin:6px 0;}

.memo-container{
    background:linear-gradient(135deg,#0a1628 0%,#0d1f45 100%);
    border:1px solid #283593;border-radius:12px;padding:22px;line-height:1.7;}

.dep-chip{display:inline-block;background:#1a237e;color:#c5cae9;
    border-radius:4px;padding:2px 8px;font-size:.75rem;margin:2px;}
.badge-sheet{background:#283593;color:#e8eaf6;border-radius:4px;
    padding:2px 7px;font-size:.72rem;font-weight:600;}
.badge-cell{background:#0d1f45;color:#9fa8da;border-radius:4px;
    padding:2px 5px;font-size:.7rem;font-family:monospace;}
hr{border-color:#1a2f5c;}
</style>
""", unsafe_allow_html=True)

# ── Load configs ──────────────────────────────────────────────────────────────
@st.cache_data
def load_registry():
    with open(ROOT / "config" / "template_registry.json", encoding="utf-8") as f:
        registry = json.load(f)
    registry.setdefault("Infrastructure", {})["Smart Meter Rollout"] = "examples/idemo.xlsx"
    return registry

@st.cache_data
def load_input_schema():
    return load_input_schema_config()

@st.cache_data
def load_driver_registry_cached():
    return load_driver_registry_config()

registry     = load_registry()
input_schema = load_input_schema()
driver_reg   = load_driver_registry_cached()

# ── Signal helper ─────────────────────────────────────────────────────────────
SIGNALS = {
    "irr_pct":           lambda v: "🟢" if v >= 14 else ("🟡" if v >= 10 else "🔴"),
    "equity_irr_pct":    lambda v: "🟢" if v >= 18 else ("🟡" if v >= 13 else "🔴"),
    "dscr_avg":          lambda v: "🟢" if v >= 1.3 else ("🟡" if v >= 1.1 else "🔴"),
    "payback_yrs":       lambda v: "🟢" if v <= 8 else ("🟡" if v <= 12 else "🔴"),
    "ebitda_margin_pct": lambda v: "🟢" if v >= 50 else ("🟡" if v >= 30 else "🔴"),
    "npv_cr":            lambda v: "🟢" if v > 0 else "🔴",
}

# ── KPI renderer ──────────────────────────────────────────────────────────────
def render_kpi_tiles(metrics: dict, prefix: str = "") -> None:
    """Render profitability + returns KPI tiles."""
    from engines.metrics_extractor import METRIC_LABELS

    st.markdown("#### 📈 Profitability")
    c1, c2, c3, c4 = st.columns(4)
    prof_map = [
        (c1, "revenue_cr",        "₹ Cr"),
        (c2, "ebitda_cr",         "₹ Cr"),
        (c3, "pat_cr",            "₹ Cr"),
        (c4, "ebitda_margin_pct", "%"),
    ]
    for col, key, sfx in prof_map:
        val = metrics.get(key)
        sig = SIGNALS.get(key, lambda v: "")(val) if val is not None else ""
        with col:
            st.markdown(f"""
            <div class="kpi-tile">
                <div class="label">{METRIC_LABELS.get(key, key)}</div>
                <div class="value">{f"{val:,.2f} {sfx}" if val is not None else "—"}</div>
                <div class="signal">{sig}</div>
            </div>""", unsafe_allow_html=True)

    st.markdown("#### 🎯 Investment Returns")
    r1, r2, r3, r4, r5 = st.columns(5)
    ret_map = [
        (r1, "irr_pct",        "%"),
        (r2, "equity_irr_pct", "%"),
        (r3, "dscr_avg",       "x"),
        (r4, "payback_yrs",    " Yrs"),
        (r5, "npv_cr",         " ₹ Cr"),
    ]
    for col, key, sfx in ret_map:
        val = metrics.get(key)
        sig = SIGNALS.get(key, lambda v: "")(val) if val is not None else ""
        with col:
            st.markdown(f"""
            <div class="kpi-tile">
                <div class="label">{METRIC_LABELS.get(key, key)}</div>
                <div class="value">{f"{val:,.2f}{sfx}" if val is not None else "—"}</div>
                <div class="signal">{sig}</div>
            </div>""", unsafe_allow_html=True)


# ── Results panel (shared by both tabs) ────────────────────────────────────────
def render_results_panel(
    model_bytes: bytes | None,
    model_name: str,
    metrics: dict | None,
    project_type: str,
    industry: str,
    assumptions: dict,
    tab_prefix: str,
) -> None:
    """Renders download, KPIs, analysis, and report sections after model generation."""
    if not model_bytes:
        return

    st.markdown('<div class="alert-success">Financial model generated successfully!</div>',
                unsafe_allow_html=True)

    dl_col, _ = st.columns([1, 2])
    with dl_col:
        st.download_button(
            "📥 Download Excel Model", data=model_bytes,
            file_name=model_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key=f"{tab_prefix}_dl_model", use_container_width=True,
        )

    st.markdown("---")
    st.markdown("""<div class="step-header">
    <h3>Financial Metrics Summary</h3>
    <p>Key performance indicators from the financial model</p></div>""",
                unsafe_allow_html=True)

    if metrics:
        render_kpi_tiles(metrics, prefix=tab_prefix)
    else:
        st.info("No metrics available — use a real Excel template with formulas for live KPIs.")

    st.markdown("<br>", unsafe_allow_html=True)
    st.markdown("---")

    # Analysis
    st.markdown("""<div class="step-header">
    <h3>AI Investment Analysis</h3>
    <p>Professional investment memo with risk assessment and recommendation</p></div>""",
                unsafe_allow_html=True)

    ana_memo_key = f"{tab_prefix}_memo"
    ana_src_key  = f"{tab_prefix}_memo_src"
    if ana_memo_key not in st.session_state: st.session_state[ana_memo_key] = None
    if ana_src_key  not in st.session_state: st.session_state[ana_src_key]  = None

    a_col, _ = st.columns([1, 2])
    with a_col:
        ana_btn = st.button("🤖 Generate Investment Analysis", key=f"{tab_prefix}_ana_btn",
                            use_container_width=True)

    if ana_btn:
        with st.spinner("Generating investment analysis..."):
            from engines.analysis_engine import generate_analysis
            memo, source = generate_analysis(metrics or {}, project_type, industry, assumptions)
            st.session_state[ana_memo_key] = memo
            st.session_state[ana_src_key]  = source

    if st.session_state.get(ana_memo_key):
        src = st.session_state[ana_src_key]
        badge = "🤖 AI-Generated (Groq)" if src == "groq" else "📐 Rule-Based Analysis"
        st.markdown(f'<div class="alert-info">{badge}</div>', unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("""<div class="step-header"><h3>Download Investment Report</h3>
        <p>Export the full analysis as a PDF or Markdown report</p></div>""",
                    unsafe_allow_html=True)

        rc1, rc2, _ = st.columns([1, 1, 1])
        from engines.report_generator import generate_pdf_report, is_pdf_available
        try:
            pdf_bytes, pdf_name = generate_pdf_report(
                st.session_state[ana_memo_key], metrics or {},
                project_type, industry,
            )
            mime = "application/pdf" if pdf_name.endswith(".pdf") else "text/markdown"
            with rc1:
                st.download_button(
                    f"📄 Download {'PDF' if pdf_name.endswith('.pdf') else 'Markdown'} Report",
                    data=pdf_bytes, file_name=pdf_name, mime=mime,
                    key=f"{tab_prefix}_dl_report", use_container_width=True,
                )
        except Exception as exc:
            st.warning(f"Report generation: {exc}")

        with rc2:
            st.download_button(
                "📝 Download Memo (Markdown)",
                data=st.session_state[ana_memo_key].encode("utf-8"),
                file_name=f"memo_{project_type.replace(' ','_').lower()}.md",
                mime="text/markdown",
                key=f"{tab_prefix}_dl_memo", use_container_width=True,
            )

        st.markdown("---")
        st.markdown("### 📋 Investment Analysis Memo")
        st.markdown(
            f'<div class="memo-container">{st.session_state[ana_memo_key]}</div>',
            unsafe_allow_html=True,
        )


# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="text-align:center;padding:10px 0 18px;">
        <div style="font-size:2rem;margin-bottom:4px;">📊</div>
        <div style="font-weight:700;font-size:1rem;color:#c5cae9;">AI Financial Model</div>
        <div style="font-size:.72rem;color:#5c6e91;">Platform v2.0</div>
    </div>""", unsafe_allow_html=True)
    st.markdown("---")
    groq_key = st.text_input("Groq API Key (optional)", type="password",
                              placeholder="gsk_...",
                              help="For AI memos. Leave blank for rule-based analysis.")
    if groq_key:
        os.environ["GROQ_API_KEY"] = groq_key
        st.markdown('<div class="alert-success">Groq API key set</div>', unsafe_allow_html=True)
    else:
        st.markdown('<div class="alert-info">Using rule-based analysis</div>', unsafe_allow_html=True)
    st.markdown("---")
    st.markdown("""
    <div style="font-size:.7rem;color:#5c6e91;text-align:center;">
    AI Financial Model Platform v2.0<br>Built with Streamlit + openpyxl<br>
    For informational use only
    </div>""", unsafe_allow_html=True)

# ── Banner ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="platform-banner">
    <h1>📊 AI Financial Model Platform</h1>
    <p>Intelligent template scanning · Dynamic form generation · AI investment analysis</p>
</div>""", unsafe_allow_html=True)

# ── Tabs ──────────────────────────────────────────────────────────────────────
tab1, tab2 = st.tabs(["🏭  Curated Templates", "📤  Upload Your Template"])

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — CURATED TEMPLATES
# ─────────────────────────────────────────────────────────────────────────────
with tab1:
    c1, c2 = st.columns(2)
    with c1:
        st.markdown("**Step 1 — Industry**")
        industries = list(registry.keys())
        sel_industry = st.selectbox("Industry Sector", industries, key="t1_industry")
    with c2:
        st.markdown("**Step 2 — Project Type**")
        proj_types = list(registry[sel_industry].keys())
        sel_project = st.selectbox("Project Type", proj_types, key="t1_project")

    tmpl_rel  = registry[sel_industry][sel_project]
    tmpl_path = ROOT / tmpl_rel

    if tmpl_path.exists():
        st.markdown('<div class="alert-success">✅ Template found and ready</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div class="alert-warning">⚠️ Template not found: {tmpl_rel}<br>'
                    'Run: <code>python scripts/create_demo_templates.py</code></div>',
                    unsafe_allow_html=True)

    st.markdown("---")
    st.markdown("""<div class="step-header"><h3>Step 3 — Enter Assumptions</h3>
    <p>Fill in project-specific financial parameters</p></div>""", unsafe_allow_html=True)

    fields = flatten_project_fields(input_schema.get(sel_project, []))
    assumptions_t1: dict[str, Any] = {}

    if isinstance(input_schema.get(sel_project), dict):
        for group_name, group_fields in input_schema.get(sel_project, {}).items():
            st.markdown(f"**{group_name}**")
            for row_start in range(0, len(group_fields), 3):
                row_fields = group_fields[row_start:row_start + 3]
                cols = st.columns(len(row_fields))
                for col_w, field in zip(cols, row_fields):
                    name    = field["name"]
                    ftype   = field.get("type", "float")
                    fmin    = field.get("min", 0)
                    fmax    = field.get("max", 1e9)
                    default = field.get("default", fmin)
                    unit    = field.get("unit", "")
                    label   = f"{field['label']} *({unit})*" if unit else field["label"]
                    with col_w:
                        if ftype == "integer":
                            val = st.number_input(label, min_value=int(fmin), max_value=int(fmax),
                                                  value=int(default), step=1, key=f"t1_{name}")
                        else:
                            step = 0.01 if fmax <= 1 else (0.1 if fmax <= 100 else 1.0)
                            val = st.number_input(label, min_value=float(fmin), max_value=float(fmax),
                                                  value=float(default), step=step,
                                                  format="%.4f" if fmax <= 1 else "%.2f",
                                                  key=f"t1_{name}")
                        assumptions_t1[name] = val
    else:
        for row_start in range(0, len(fields), 3):
            row_fields = fields[row_start:row_start + 3]
            cols = st.columns(len(row_fields))
            for col_w, field in zip(cols, row_fields):
                name    = field["name"]
                ftype   = field.get("type", "float")
                fmin    = field.get("min", 0)
                fmax    = field.get("max", 1e9)
                default = field.get("default", fmin)
                unit    = field.get("unit", "")
                label   = f"{field['label']} *({unit})*" if unit else field["label"]
                with col_w:
                    if ftype == "integer":
                        val = st.number_input(label, min_value=int(fmin), max_value=int(fmax),
                                              value=int(default), step=1, key=f"t1_{name}")
                    else:
                        step = 0.01 if fmax <= 1 else (0.1 if fmax <= 100 else 1.0)
                        val = st.number_input(label, min_value=float(fmin), max_value=float(fmax),
                                              value=float(default), step=step,
                                              format="%.4f" if fmax <= 1 else "%.2f",
                                              key=f"t1_{name}")
                    assumptions_t1[name] = val

    from engines.validation_engine import validate, coerce
    t1_errors = validate(sel_project, assumptions_t1)
    if t1_errors:
        for msg in t1_errors.values():
            st.error(msg)
    else:
        assumptions_t1 = coerce(sel_project, assumptions_t1)

    if driver_reg.get(sel_project):
        with st.expander("🔍 Driver Cell Mapping Preview", expanded=False):
            from engines.assumption_mapper import describe_mapping
            import pandas as pd
            mappings = describe_mapping(sel_project, assumptions_t1)
            if mappings:
                df = pd.DataFrame(mappings)[["sheet","cell","label","value"]]
                df.columns = ["Sheet","Cell","Driver","Value"]
                st.dataframe(df, use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("""<div class="step-header"><h3>Step 4 — Generate Financial Model</h3>
    <p>Writes assumptions into the template, preserving all formulas</p></div>""",
                unsafe_allow_html=True)

    # State
    for k in ("t1_model_bytes","t1_model_path","t1_metrics"):
        if k not in st.session_state: st.session_state[k] = None

    g_col, _ = st.columns([1, 2])
    with g_col:
        t1_gen = st.button("⚡ Generate Financial Model", key="t1_gen_btn", use_container_width=True)

    if t1_gen and not t1_errors:
        if not tmpl_path.exists():
            st.error("Template file not found. Run: python scripts/create_demo_templates.py")
        else:
            with st.spinner("Generating model..."):
                from engines.assumption_mapper import map_assumptions
                from engines.template_writer   import write_model
                from engines.metrics_extractor import simulate_metrics
                try:
                    write_map = map_assumptions(sel_project, assumptions_t1)
                    out_path  = write_model(tmpl_path, write_map, sel_project, sel_industry)
                    with open(out_path, "rb") as f:
                        st.session_state.t1_model_bytes = f.read()
                    st.session_state.t1_model_path  = out_path
                    st.session_state.t1_metrics     = simulate_metrics(assumptions_t1, sel_project)
                    st.session_state.t1_memo        = None
                    st.session_state.t1_memo_src    = None
                except Exception as exc:
                    st.error(f"Generation failed: {exc}")
                    st.exception(exc)

    render_results_panel(
        model_bytes  = st.session_state.get("t1_model_bytes"),
        model_name   = (st.session_state.t1_model_path.name
                        if st.session_state.get("t1_model_path") else "model.xlsx"),
        metrics      = st.session_state.get("t1_metrics"),
        project_type = sel_project,
        industry     = sel_industry,
        assumptions  = assumptions_t1,
        tab_prefix   = "t1",
    )

# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — UPLOAD TEMPLATE
# ─────────────────────────────────────────────────────────────────────────────
with tab2:

    # ── State initialisation ────────────────────────────────────────────────
    for k in ("t2_scan_result","t2_driver_reg","t2_dep_map",
              "t2_model_bytes","t2_model_path","t2_metrics",
              "t2_memo","t2_memo_src","t2_tmpl_path","t2_industry","t2_project"):
        if k not in st.session_state:
            st.session_state[k] = None

    # ── Step 1: Upload ──────────────────────────────────────────────────────
    st.markdown("""<div class="step-header"><h3>Step 1 — Upload Excel Template</h3>
    <p>Upload any .xlsx financial model. The scanner will detect its driver inputs automatically.</p>
    </div>""", unsafe_allow_html=True)

    col_up, col_meta = st.columns([1, 1])
    with col_up:
        uploaded = st.file_uploader(
            "Upload your Excel template (.xlsx)",
            type=["xlsx"],
            key="t2_uploader",
            help="Supports any .xlsx workbook. Driver cells are auto-detected in NTBA, CAPEX, TBA sheets.",
        )

    with col_meta:
        if uploaded:
            st.markdown(f"""
            <div class="scanner-box">
                <div style="font-weight:600;color:#c5cae9;margin-bottom:8px;">📁 File Details</div>
                <div style="font-size:.85rem;color:#9fa8da;">
                    <b>Name:</b> {uploaded.name}<br>
                    <b>Size:</b> {uploaded.size / 1024:.1f} KB<br>
                    <b>Type:</b> Excel Workbook (.xlsx)
                </div>
            </div>""", unsafe_allow_html=True)

    # ── Step 2: Configure driver sheets ────────────────────────────────────
    if uploaded:
        st.markdown("---")
        st.markdown("""<div class="step-header"><h3>Step 2 — Configure Scanner</h3>
        <p>Specify which sheets contain driver (input) cells, and set project metadata</p></div>""",
                    unsafe_allow_html=True)

        mc1, mc2, mc3 = st.columns(3)
        with mc1:
            driver_sheets_input = st.text_input(
                "Driver sheets (comma-separated)",
                value="NTBA, CAPEX, TBA",
                key="t2_driver_sheets",
                help="Sheets to scan for numeric input cells",
            )
            driver_sheets = [s.strip() for s in driver_sheets_input.split(",") if s.strip()]

        with mc2:
            t2_industry = st.text_input("Industry / Sector", value="Custom",
                                         placeholder="Energy, Infrastructure …", key="t2_industry_inp")
        with mc3:
            t2_project = st.text_input("Project Type", value="Custom Project",
                                        placeholder="Solar, Highway …", key="t2_project_inp")

        # ── Scan button ──────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("""<div class="step-header"><h3>Step 3 — Run Template Scanner</h3>
        <p>Scans all sheets, detects formula vs input cells, labels driver cells, maps dependencies</p>
        </div>""", unsafe_allow_html=True)

        scan_col, _ = st.columns([1, 2])
        with scan_col:
            scan_btn = st.button("🔍 Run Template Scanner", key="t2_scan_btn", use_container_width=True)

        if scan_btn:
            with st.spinner("Scanning template — this may take a moment for large workbooks..."):
                from engines.template_scanner  import scan_template
                from engines.label_detector    import detect_labels
                from engines.dependency_mapper import map_dependencies

                # Save upload to a temp file
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp.write(uploaded.getvalue())
                    tmp_path = Path(tmp.name)

                # Per-session config dir to avoid collision
                session_config = ROOT / "config" / "sessions" / uploaded.name.replace(".xlsx","")
                session_config.mkdir(parents=True, exist_ok=True)

                try:
                    # Run full pipeline
                    scan_result = scan_template(
                        tmp_path,
                        driver_sheets=driver_sheets,
                        save_to_config=True,
                        config_dir=session_config,
                    )

                    driver_registry_scanned = detect_labels(
                        tmp_path,
                        driver_candidates=scan_result["driver_candidates"],
                        save_to_config=True,
                        config_dir=session_config,
                    )

                    dep_map = map_dependencies(
                        tmp_path,
                        save_to_config=True,
                        config_dir=session_config,
                    )

                    # Persist to session state
                    st.session_state.t2_scan_result  = scan_result
                    st.session_state.t2_driver_reg   = driver_registry_scanned
                    st.session_state.t2_dep_map      = dep_map
                    st.session_state.t2_tmpl_path    = tmp_path
                    st.session_state.t2_industry     = t2_industry
                    st.session_state.t2_project      = t2_project
                    # Reset downstream state
                    for k in ("t2_model_bytes","t2_model_path","t2_metrics","t2_memo","t2_memo_src"):
                        st.session_state[k] = None

                except Exception as exc:
                    st.error(f"Scanner error: {exc}")
                    st.exception(exc)

    # ── Display scan results ─────────────────────────────────────────────────
    if st.session_state.get("t2_scan_result"):
        scan  = st.session_state.t2_scan_result
        dreg  = st.session_state.t2_driver_reg
        deps  = st.session_state.t2_dep_map
        summ  = scan["summary"]

        st.markdown("---")
        st.markdown("""<div class="step-header"><h3>Scanner Results</h3>
        <p>Template structure, formulas, driver candidates, and sheet dependencies</p></div>""",
                    unsafe_allow_html=True)

        # Summary stats row
        s1, s2, s3, s4 = st.columns(4)
        for col, num, lbl in [
            (s1, summ["sheet_count"],              "Sheets Detected"),
            (s2, summ["total_formula_cells"],      "Formula Cells"),
            (s3, summ["total_numeric_cells"],      "Numeric Cells"),
            (s4, summ["total_driver_candidates"],  "Driver Candidates"),
        ]:
            col.markdown(f"""
            <div class="kpi-tile">
                <div class="label">{lbl}</div>
                <div class="value">{num:,}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # Sheet structure table
        with st.expander("📋 Sheet Structure", expanded=True):
            import pandas as pd
            ss = scan["sheet_structure"]
            rows = []
            for name, info in ss.items():
                rows.append({
                    "Sheet":          name,
                    "Type":           info["classification"].capitalize(),
                    "Formula Cells":  info["formula_cell_count"],
                    "Numeric Cells":  info["numeric_cell_count"],
                    "Driver Sheet":   "✅" if info["is_driver_sheet"] else "—",
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        # Dependency graph
        with st.expander("🔗 Sheet Dependencies", expanded=False):
            non_empty = {k: v for k, v in deps.items() if v}
            if non_empty:
                for sheet, refs in sorted(non_empty.items()):
                    chips = "".join(f'<span class="dep-chip">{r}</span>' for r in refs)
                    st.markdown(
                        f'<div style="padding:4px 0;font-size:.85rem;">'
                        f'<b style="color:#c5cae9;">{sheet}</b> '
                        f'<span style="color:#5c6e91;margin:0 6px;">reads from</span>'
                        f'{chips}</div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.info("No cross-sheet formula references detected.")

        # ── Step 4: Detected drivers ────────────────────────────────────────
        st.markdown("---")
        st.markdown("""<div class="step-header"><h3>Step 4 — Detected Driver Inputs</h3>
        <p>These cells were identified as numeric inputs (non-formula) in the specified driver sheets</p>
        </div>""", unsafe_allow_html=True)

        total_drivers = sum(len(v) for v in dreg.values())

        if total_drivers == 0:
            st.markdown(f"""
            <div class="alert-warning">
            No driver cells found in sheets: {', '.join(driver_sheets)}<br>
            This usually means: the template was saved with <b>data_only</b> values computed,
            OR those sheets don't exist in this workbook.<br>
            Try adjusting the <b>Driver sheets</b> field above.
            </div>""", unsafe_allow_html=True)
        else:
            st.markdown(f'<div class="alert-success">Found {total_drivers} driver cells across '
                        f'{len(dreg)} sheet(s)</div>', unsafe_allow_html=True)

            for sheet_name, entries in dreg.items():
                src_counts = {"left": 0, "above": 0, "fallback": 0}
                for e in entries:
                    src_counts[e.get("source", "fallback")] += 1
                badges = " · ".join(
                    f"{v} {k}" for k, v in src_counts.items() if v
                )
                st.markdown(
                    f'<span style="font-weight:600;color:#c5cae9;">{sheet_name}</span> '
                    f'<span style="color:#5c6e91;font-size:.8rem;">({len(entries)} drivers · detected: {badges})</span>',
                    unsafe_allow_html=True,
                )
                rows = []
                for e in entries:
                    src = e.get("source", "?")
                    src_icon = {"left": "⬅️ left cell", "above": "⬆️ above cell",
                                "fallback": "⚠️ fallback"}.get(src, src)
                    rows.append({
                        "Cell":          e["cell"],
                        "Detected Label": e["label"],
                        "Full Label (w/ units)": e.get("raw_label", e["label"]),
                        "Default":       e.get("value", ""),
                        "Detected From": src_icon,
                    })
                if rows:
                    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

        # ── Step 5: Dynamic assumption form ────────────────────────────────
        if total_drivers > 0:
            st.markdown("---")
            st.markdown("""<div class="step-header"><h3>Step 5 — Enter Assumptions</h3>
            <p>Form auto-generated from template labels. Clean names shown; hover for unit details.</p>
            </div>""", unsafe_allow_html=True)

            assumptions_t2: dict[str, Any] = {}

            # Render inputs grouped by sheet with section headers
            for sheet_name, entries in dreg.items():
                if not entries:
                    continue
                st.markdown(
                    f'<p style="font-weight:600;color:#9fa8da;font-size:.85rem;'
                    f'text-transform:uppercase;letter-spacing:.06em;margin:14px 0 6px;">{sheet_name} Inputs</p>',
                    unsafe_allow_html=True,
                )
                for row_start in range(0, len(entries), 3):
                    row_items = entries[row_start:row_start + 3]
                    cols = st.columns(len(row_items))
                    for col_w, drv in zip(cols, row_items):
                        cell      = drv["cell"]
                        label     = drv["label"]           # clean label (units stripped)
                        raw_label = drv.get("raw_label", label)  # original with units → help text
                        default   = float(drv.get("value") or 0)
                        key       = f"t2_{sheet_name}_{cell}"
                        help_txt  = raw_label if raw_label != label else None
                        with col_w:
                            # ── Smart numeric type heuristics ──────────────
                            # Ratio/proportion (0 < v ≤ 1)  → 4 decimal places
                            if 0 < default <= 1.0:
                                val = st.number_input(
                                    label, value=default, step=0.01,
                                    format="%.4f", min_value=0.0, max_value=1.0,
                                    key=key, help=help_txt,
                                )
                            # Percentage (1 < v ≤ 100)
                            elif 1 < default <= 100:
                                val = st.number_input(
                                    label, value=default, step=0.1,
                                    format="%.2f", min_value=0.0, max_value=100.0,
                                    key=key, help=help_txt,
                                )
                            # Large integer (> 1000 and no decimal part)
                            elif default > 1000 and default == int(default):
                                val = st.number_input(
                                    label, value=int(default), step=1000,
                                    min_value=0, key=key, help=help_txt,
                                )
                            # General float (years, km, MW, Rs Cr, etc.)
                            else:
                                step = 0.5 if default < 10 else (5.0 if default < 100 else 10.0)
                                val = st.number_input(
                                    label, value=default, step=step,
                                    format="%.2f", min_value=0.0,
                                    key=key, help=help_txt,
                                )
                        assumptions_t2[f"{sheet_name}::{cell}"] = val

            # ── Step 6: Generate model ──────────────────────────────────
            st.markdown("---")
            st.markdown("""<div class="step-header"><h3>Step 6 — Generate Financial Model</h3>
            <p>Writes your assumptions into the template while preserving all existing formulas</p>
            </div>""", unsafe_allow_html=True)

            gen_col2, _ = st.columns([1, 2])
            with gen_col2:
                t2_gen = st.button("⚡ Generate Financial Model", key="t2_gen_btn",
                                   use_container_width=True)

            if t2_gen:
                tmpl_path_t2 = st.session_state.get("t2_tmpl_path")
                if not tmpl_path_t2 or not Path(tmpl_path_t2).exists():
                    st.error("Template file not found. Please re-upload and re-scan.")
                else:
                    with st.spinner("Generating financial model..."):
                        import openpyxl, shutil
                        from datetime import datetime
                        from engines.template_writer import ALLOWED_SHEETS, OUTPUT_DIR

                        try:
                            # Build write_map: {sheet: {cell: value}}
                            write_map: dict[str, dict[str, Any]] = {}
                            for key_str, val in assumptions_t2.items():
                                sheet_name, cell_addr = key_str.split("::", 1)
                                if sheet_name not in write_map:
                                    write_map[sheet_name] = {}
                                write_map[sheet_name][cell_addr] = val

                            # Copy + write
                            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
                            safe = (st.session_state.t2_project or "custom").replace(" ","_").lower()
                            out_name = f"custom_{safe}_{ts}.xlsx"
                            out_path = OUTPUT_DIR / out_name
                            OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
                            shutil.copy2(tmpl_path_t2, out_path)

                            wb = openpyxl.load_workbook(str(out_path))
                            for sname, cell_map in write_map.items():
                                if sname not in wb.sheetnames:
                                    continue
                                ws = wb[sname]
                                for caddr, cval in cell_map.items():
                                    c = ws[caddr]
                                    existing = c.value
                                    if isinstance(existing, str) and existing.strip().startswith("="):
                                        continue  # never overwrite formulas
                                    c.value = cval
                            wb.save(str(out_path))
                            wb.close()

                            with open(out_path, "rb") as f:
                                st.session_state.t2_model_bytes = f.read()
                            st.session_state.t2_model_path = out_path

                            # Simulate basic metrics
                            from engines.metrics_extractor import simulate_metrics
                            proj = st.session_state.t2_project or "Custom Project"
                            met  = simulate_metrics(assumptions_t2, proj)
                            # met will likely be zeroes for custom — that's fine
                            st.session_state.t2_metrics = met

                        except Exception as exc:
                            st.error(f"Generation failed: {exc}")
                            st.exception(exc)

            render_results_panel(
                model_bytes  = st.session_state.get("t2_model_bytes"),
                model_name   = (st.session_state.t2_model_path.name
                                if st.session_state.get("t2_model_path") else "generated_model.xlsx"),
                metrics      = st.session_state.get("t2_metrics"),
                project_type = st.session_state.get("t2_project") or "Custom Project",
                industry     = st.session_state.get("t2_industry") or "Custom",
                assumptions  = assumptions_t2,
                tab_prefix   = "t2",
            )

    # ── "How it works" shown before upload ──────────────────────────────────
    if not uploaded:
        st.markdown("---")
        st.markdown("### How the Template Scanner Works")
        steps = [
            ("📤", "Upload Excel",     "Any .xlsx financial model template"),
            ("🔍", "Auto-Scan",        "Detects formula vs input cells in NTBA / CAPEX / TBA"),
            ("🏷️",  "Label Detection",  "Reads adjacent cells to name each driver"),
            ("🔗", "Dependency Map",   "Parses formulas to graph sheet relationships"),
            ("📝", "Dynamic Form",     "Auto-generates assumption input fields"),
            ("⚡", "Generate Model",   "Writes values into template, preserves formulas"),
        ]
        cols = st.columns(3)
        for i, (icon, title, desc) in enumerate(steps):
            with cols[i % 3]:
                st.markdown(f"""
                <div class="metric-card">
                    <div style="font-size:1.5rem;margin-bottom:6px;">{icon}</div>
                    <div style="font-weight:600;color:#e8eaf6;margin-bottom:4px;">{title}</div>
                    <div style="font-size:.8rem;color:#9fa8da;">{desc}</div>
                </div>""", unsafe_allow_html=True)
